import os
import sys
import json
import base64
import secrets
import mimetypes
from datetime import datetime, date, timedelta
from io import BytesIO

from flask import Flask, request, jsonify, send_from_directory, redirect, render_template, flash, url_for, send_file
from flask_cors import CORS
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BILLING_DIR = os.path.join(BASE_DIR, 'Billing system')
ENCRYPTION_KEY_FILE = os.path.join(BASE_DIR, '.server_key')
IS_VERCEL = os.environ.get('VERCEL', '') == '1'

DATABASE_URL = os.environ.get('DATABASE_URL', '')
SUPABASE_URL = os.environ.get('SUPABASE_URL', 'https://xpkaoqwywhvliwbbuwoh.supabase.co')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY', 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Inhwa2FvcXd5d2h2bGl3YmJ1d29oIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODQ2NTMxMzcsImV4cCI6MjEwMDIyOTEzN30.HEG6AD625_UQ8tD8CH-zHMKWOrPPFb-tk_iJCA1i3X4')
SUPABASE_STORAGE_BUCKET = os.environ.get('SUPABASE_STORAGE_BUCKET', 'uploads')
TURSO_DATABASE_URL = os.environ.get('TURSO_DATABASE_URL', '')
TURSO_AUTH_TOKEN = os.environ.get('TURSO_AUTH_TOKEN', '')

if not IS_VERCEL:
    os.makedirs(os.path.join(BILLING_DIR, 'instance'), exist_ok=True)

app = Flask(__name__, static_folder=os.path.join(BILLING_DIR, 'static'), template_folder=os.path.join(BILLING_DIR, 'templates'))
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

if DATABASE_URL:
    app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    pool_opts = {'pool_pre_ping': True, 'pool_recycle': 300}
    if 'postgresql' in DATABASE_URL:
        pool_opts['connect_args'] = {'sslmode': 'require'}
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = pool_opts
elif TURSO_DATABASE_URL and TURSO_AUTH_TOKEN:
    try:
        import libsql_experimental as libsql
        def _create_turso_connection():
            return libsql.connect('/tmp/turso.db', sync_url=TURSO_DATABASE_URL, auth_token=TURSO_AUTH_TOKEN)
        app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:////tmp/turso.db'
        app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
        app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'creator': _create_turso_connection}
    except ImportError:
        print("[DB] libsql-experimental not available, using /tmp SQLite")
        app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:////tmp/billing.db'
        app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
        app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {}
elif IS_VERCEL:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:////tmp/billing.db'
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {}
else:
    billing_db = os.path.join(BILLING_DIR, 'instance', 'billing.db').replace('\\', '/')
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + billing_db
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {}
CORS(app)

def get_or_create_server_key():
    env_key = os.environ.get('ENCRYPTION_KEY', '')
    if env_key:
        return env_key
    if os.path.exists(ENCRYPTION_KEY_FILE):
        with open(ENCRYPTION_KEY_FILE, 'r') as f:
            return f.read().strip()
    key = Fernet.generate_key().decode()
    if not IS_VERCEL:
        with open(ENCRYPTION_KEY_FILE, 'w') as f:
            f.write(key)
    return key

SERVER_KEY = get_or_create_server_key()
cipher_suite = Fernet(SERVER_KEY.encode())

def derive_user_key(password, salt=b'iacss_server_salt_v1'):
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=100000)
    return base64.urlsafe_b64encode(kdf.derive(password.encode()))

def encrypt_data(data, user_key=None):
    if isinstance(data, str):
        data = data.encode()
    if user_key:
        return Fernet(derive_user_key(user_key)).encrypt(data).decode()
    return cipher_suite.encrypt(data).decode()

def decrypt_data(encrypted, user_key=None):
    if isinstance(encrypted, str):
        encrypted = encrypted.encode()
    if user_key:
        return Fernet(derive_user_key(user_key)).decrypt(encrypted).decode()
    return cipher_suite.decrypt(encrypted).decode()

attendance_log_file = os.path.join(BASE_DIR, '.attendance_log.json') if not IS_VERCEL else None

def log_attendance(username, system, role='user'):
    record = {
        'id': int(datetime.now().timestamp() * 1000),
        'username': username, 'system': system, 'role': role,
        'timestamp': datetime.now().isoformat(),
        'date': datetime.now().strftime('%Y-%m-%d'),
        'time': datetime.now().strftime('%H:%M:%S'),
        'ip': request.remote_addr
    }
    if IS_VERCEL or not attendance_log_file:
        return record
    records = []
    if os.path.exists(attendance_log_file):
        try:
            with open(attendance_log_file, 'r') as f:
                records = json.loads(decrypt_data(f.read()))
        except Exception:
            records = []
    records.insert(0, record)
    if len(records) > 1000:
        records = records[:1000]
    encrypted = encrypt_data(json.dumps(records))
    with open(attendance_log_file, 'w') as f:
        f.write(encrypted)
    return record


from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash

db = SQLAlchemy()

class CompanySettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(200), default='Indian A/C Sales and Services')
    tagline = db.Column(db.String(300), default='Home Appliances & Furnitures')
    address = db.Column(db.Text, default='')
    city = db.Column(db.String(100), default='')
    state = db.Column(db.String(100), default='')
    pincode = db.Column(db.String(10), default='')
    phone = db.Column(db.String(20), default='')
    email = db.Column(db.String(120), default='')
    website = db.Column(db.String(200), default='')
    gstin = db.Column(db.String(20), default='')
    pan = db.Column(db.String(20), default='')
    bank_name = db.Column(db.String(100), default='')
    bank_account = db.Column(db.String(30), default='')
    bank_ifsc = db.Column(db.String(15), default='')
    logo_filename = db.Column(db.String(100), default='logo.svg')
    billing_address = db.Column(db.Text, default='')
    shipping_address = db.Column(db.Text, default='')
    place_of_supply = db.Column(db.String(100), default='')

class BUser(UserMixin, db.Model):
    __tablename__ = 'b_user'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    full_name = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='staff')
    is_active_user = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120))
    phone = db.Column(db.String(20))
    address = db.Column(db.Text)
    shipping_address = db.Column(db.Text)
    state = db.Column(db.String(100))
    state_code = db.Column(db.String(5))
    gst_number = db.Column(db.String(20))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    invoices = db.relationship('Invoice', backref='customer', lazy=True)

class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    description = db.Column(db.Text)
    price = db.Column(db.Float, nullable=False)
    price_with_gst = db.Column(db.Float, default=0.0)
    stock = db.Column(db.Integer, default=0)
    gst_rate = db.Column(db.Float, default=18.0)
    hsn_code = db.Column(db.String(20))
    brand = db.Column(db.String(100), default='')
    category = db.Column(db.String(100), default='')
    unit = db.Column(db.String(20), default='Pcs')
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Invoice(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    invoice_number = db.Column(db.String(20), unique=True, nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('b_user.id'), nullable=False)
    date_created = db.Column(db.DateTime, default=datetime.utcnow)
    subtotal = db.Column(db.Float, default=0.0)
    cgst = db.Column(db.Float, default=0.0)
    sgst = db.Column(db.Float, default=0.0)
    igst = db.Column(db.Float, default=0.0)
    total_tax = db.Column(db.Float, default=0.0)
    total_amount = db.Column(db.Float, default=0.0)
    discount = db.Column(db.Float, default=0.0)
    round_off = db.Column(db.Float, default=0.0)
    grand_total = db.Column(db.Float, default=0.0)
    status = db.Column(db.String(20), default='unpaid')
    place_of_supply = db.Column(db.String(50))
    due_date = db.Column(db.String(20))
    reverse_charge = db.Column(db.Boolean, default=False)
    user = db.relationship('BUser', backref='invoices_created')
    items = db.relationship('InvoiceItem', backref='invoice')
    payments = db.relationship('Payment', backref='invoice')

class InvoiceItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    invoice_id = db.Column(db.Integer, db.ForeignKey('invoice.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    unit_price = db.Column(db.Float, nullable=False)
    discount = db.Column(db.Float, default=0.0)
    taxable_value = db.Column(db.Float, default=0.0)
    gst_rate = db.Column(db.Float, default=0.0)
    gst_amount = db.Column(db.Float, default=0.0)
    igst = db.Column(db.Float, default=0.0)
    total = db.Column(db.Float, nullable=False)
    star_rating = db.Column(db.String(10), default='')
    model_no = db.Column(db.String(100), default='')
    series_no = db.Column(db.String(100), default='')
    item_category = db.Column(db.String(100), default='')
    ton_size = db.Column(db.String(50), default='')
    product = db.relationship('Product')

class Payment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    invoice_id = db.Column(db.Integer, db.ForeignKey('invoice.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    payment_date = db.Column(db.DateTime, default=datetime.utcnow)
    payment_method = db.Column(db.String(20))
    reference = db.Column(db.String(100))

class BEmployee(db.Model):
    __tablename__ = 'b_employee'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('b_user.id'))
    employee_id = db.Column(db.String(20), unique=True, nullable=False)
    designation = db.Column(db.String(50))
    department = db.Column(db.String(50))
    salary = db.Column(db.Float, default=0.0)
    join_date = db.Column(db.Date, default=date.today)
    phone = db.Column(db.String(20))
    address = db.Column(db.Text)
    user = db.relationship('BUser', backref='employee_profile')
    attendances = db.relationship('BAttendance', backref='employee')

class BAttendance(db.Model):
    __tablename__ = 'b_attendance'
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('b_employee.id'), nullable=False)
    date = db.Column(db.Date, default=date.today)
    check_in = db.Column(db.DateTime)
    check_out = db.Column(db.DateTime)
    status = db.Column(db.String(20), default='present')
    notes = db.Column(db.Text)

db.init_app(app)

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message_category = 'warning'

@login_manager.user_loader
def load_user(user_id):
    return BUser.query.get(int(user_id))

@app.context_processor
def inject_now():
    company = CompanySettings.query.first()
    if not company:
        company = CompanySettings()
    logo_path = 'uploads/' + (company.logo_filename if company and company.logo_filename else 'logo.svg')
    return dict(now=datetime.now(), company=company, logo_path=logo_path)

@app.route('/')
@app.route('/index.html')
def index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/secure-api.js')
def serve_secure_api():
    return send_from_directory(BASE_DIR, 'secure-api.js')

@app.route('/logo.png')
def serve_logo():
    return send_from_directory(BASE_DIR, 'logo.png')

@app.route('/auth.js')
def serve_auth_js():
    return send_from_directory(BASE_DIR, 'auth.js')

@app.route('/dashboard')
@app.route('/dashboard.html')
def serve_dashboard():
    return send_from_directory(BASE_DIR, 'dashboard.html')

@app.route('/attendance-log.html')
def serve_attendance_log():
    return send_from_directory(BASE_DIR, 'attendance-log.html')

@app.route('/details.html')
def serve_details():
    return send_from_directory(BASE_DIR, 'details.html')

@app.route('/attendance')
@app.route('/attendance/')
def attendance_root():
    return send_from_directory(os.path.join(BASE_DIR, 'attendance'), 'index.html')

@app.route('/attendance/admin')
def attendance_admin():
    return send_from_directory(os.path.join(BASE_DIR, 'attendance'), 'admin.html')

@app.route('/attendance/leave')
def attendance_leave():
    return send_from_directory(os.path.join(BASE_DIR, 'attendance'), 'leave.html')

@app.route('/attendance/<path:filename>')
def serve_attendance(filename):
    return send_from_directory(os.path.join(BASE_DIR, 'attendance'), filename)

@app.route('/complaints')
@app.route('/complaints/')
def complaints_root():
    return send_from_directory(os.path.join(BASE_DIR, 'complaints'), 'index.html')

@app.route('/complaints/<path:filename>')
def serve_complaints(filename):
    return send_from_directory(os.path.join(BASE_DIR, 'complaints'), filename)

@app.route('/employees')
@app.route('/employees/')
def employees_root():
    return send_from_directory(os.path.join(BASE_DIR, 'employees'), 'index.html')

@app.route('/employees/<path:filename>')
def serve_employees(filename):
    return send_from_directory(os.path.join(BASE_DIR, 'employees'), filename)


@app.route('/billing')
@app.route('/billing/')
def billing_root():
    if current_user.is_authenticated:
        return redirect('/billing/dashboard')
    return redirect('/')

@app.route('/billing/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect('/billing/dashboard')
    return redirect('/')

@app.route('/billing/logout')
@login_required
def logout():
    logout_user()
    return '''<html><body><script>
sessionStorage.removeItem('iacss_unified_session');
sessionStorage.removeItem('iacss_logged_in');
sessionStorage.removeItem('iacss_user_name');
sessionStorage.removeItem('iacss_login_id');
sessionStorage.removeItem('iacss_role');
window.location.href='/';
</script></body></html>'''

@app.route('/billing/sso')
def billing_sso():
    username = request.args.get('user', '')
    role = request.args.get('role', '')
    if not username:
        return redirect('/')
    user = BUser.query.filter_by(username=username).first()
    if not user:
        user = BUser(username=username, password=generate_password_hash('sso_' + secrets.token_hex(8)), full_name=username.title(), role='admin' if role == 'admin' else 'staff')
        db.session.add(user)
        db.session.flush()
        emp = BEmployee(user_id=user.id, employee_id=f'EMP-{BUser.query.count():05d}', designation=('Admin' if role == 'admin' else 'Staff'))
        db.session.add(emp)
        db.session.commit()
    login_user(user)
    safe_user = username.replace("'", "\\'")
    safe_role = role.replace("'", "\\'")
    return f'''<html><body><script>
var s=sessionStorage;
s.setItem('iacss_unified_session',JSON.stringify({{username:'{safe_user}',role:'{safe_role}',fullName:'{safe_user}',loginTime:new Date().toISOString()}}));
s.setItem('iacss_logged_in','true');
s.setItem('iacss_user_name','{safe_user}');
s.setItem('iacss_login_id','INDIAN A/C');
s.setItem('iacss_role','{safe_role}');
window.location.href='/billing/dashboard';
</script></body></html>'''

@app.route('/billing/dashboard')
@login_required
def dashboard():
    today = date.today()
    month_start = today.replace(day=1)
    total_customers = Customer.query.count()
    total_products = Product.query.filter_by(is_active=True).count()
    total_invoices = Invoice.query.count()
    monthly_revenue = db.session.query(db.func.sum(Invoice.grand_total)).filter(Invoice.date_created >= datetime.combine(month_start, datetime.min.time()), Invoice.status == 'paid').scalar() or 0
    pending_amount = db.session.query(db.func.sum(Invoice.grand_total)).filter(Invoice.status == 'unpaid').scalar() or 0
    total_employees = BEmployee.query.count()
    low_stock_products = Product.query.filter(Product.stock < 5, Product.is_active == True).order_by(Product.stock.asc()).all()
    recent_invoices = Invoice.query.order_by(Invoice.date_created.desc()).limit(5).all()

    last_6_months = []
    for i in range(5, -1, -1):
        d = today - timedelta(days=30 * i)
        m_start = d.replace(day=1)
        if d.month == 12:
            m_end = date(d.year + 1, 1, 1)
        else:
            m_end = date(d.year, d.month + 1, 1)
        rev = db.session.query(db.func.sum(Invoice.grand_total)).filter(
            Invoice.date_created >= datetime.combine(m_start, datetime.min.time()),
            Invoice.date_created < datetime.combine(m_end, datetime.min.time()),
            Invoice.status == 'paid'
        ).scalar() or 0
        last_6_months.append({'month': m_start.strftime('%b %Y'), 'revenue': float(rev)})

    return render_template('dashboard.html',
        total_customers=total_customers, total_products=total_products,
        total_invoices=total_invoices, monthly_revenue=monthly_revenue,
        pending_amount=pending_amount, total_employees=total_employees,
        low_stock_products=low_stock_products, recent_invoices=recent_invoices,
        revenue_data=last_6_months)

@app.route('/billing/customers')
@login_required
def customers():
    search = request.args.get('search', '').strip()
    q = Customer.query
    if search:
        q = q.filter(Customer.name.ilike(f'%{search}%') | Customer.phone.ilike(f'%{search}%') | Customer.email.ilike(f'%{search}%'))
    customers_list = q.order_by(Customer.name).all()

    customer_data = []
    for c in customers_list:
        invs = Invoice.query.filter_by(customer_id=c.id).all()
        total_purchased = sum(inv.grand_total for inv in invs)
        total_paid = sum(p.amount for inv in invs for p in inv.payments)
        total_due = total_purchased - total_paid
        open_payments = []
        for inv in invs:
            paid = sum(p.amount for p in inv.payments)
            due = inv.grand_total - paid
            if due > 0.01:
                open_payments.append({'invoice': inv, 'due': due})
        products_bought = []
        for inv in invs:
            for item in inv.items:
                existing = next((pb for pb in products_bought if pb['id'] == item.product_id), None)
                if existing:
                    existing['qty'] += item.quantity
                    existing['total'] += item.total
                else:
                    products_bought.append({'id': item.product_id, 'name': item.product.name, 'qty': item.quantity, 'total': item.total})
        customer_data.append({
            'customer': c,
            'invoice_count': len(invs),
            'total_purchased': total_purchased,
            'total_due': total_due,
            'open_payments': open_payments,
            'products_bought': products_bought
        })

    return render_template('customers.html', customer_data=customer_data, search=search)

@app.route('/billing/customers/add', methods=['POST'])
@login_required
def add_customer():
    c = Customer(
        name=request.form.get('name', '').strip(),
        email=request.form.get('email', '').strip(),
        phone=request.form.get('phone', '').strip(),
        address=request.form.get('address', '').strip(),
        state=request.form.get('state', '').strip(),
        gst_number=request.form.get('gst_number', '').strip()
    )
    db.session.add(c)
    db.session.commit()
    flash('Customer added!', 'success')
    return redirect('/billing/customers')

@app.route('/billing/customers/edit/<int:id>', methods=['POST'])
@login_required
def edit_customer(id):
    c = Customer.query.get_or_404(id)
    c.name = request.form.get('name', c.name)
    c.email = request.form.get('email', c.email)
    c.phone = request.form.get('phone', c.phone)
    c.address = request.form.get('address', c.address)
    c.state = request.form.get('state', c.state)
    c.gst_number = request.form.get('gst_number', c.gst_number)
    db.session.commit()
    flash('Customer updated!', 'success')
    return redirect('/billing/customers')

@app.route('/billing/customers/delete/<int:id>', methods=['POST'])
@login_required
def delete_customer(id):
    if current_user.role != 'admin':
        flash('Admin only.', 'danger')
        return redirect('/billing/customers')
    c = Customer.query.get_or_404(id)
    for inv in c.invoices:
        for item in inv.items:
            prod = Product.query.get(item.product_id)
            if prod:
                prod.stock += item.quantity
        Payment.query.filter_by(invoice_id=inv.id).delete()
        InvoiceItem.query.filter_by(invoice_id=inv.id).delete()
        Invoice.query.filter_by(id=inv.id).delete()
    db.session.delete(c)
    db.session.commit()
    flash('Customer deleted!', 'success')
    return redirect('/billing/customers')

@app.route('/billing/products')
@login_required
def products():
    search = request.args.get('search', '').strip()
    q = Product.query.filter_by(is_active=True)
    if search:
        q = q.filter(Product.name.ilike(f'%{search}%'))
    products_list = q.order_by(Product.name).all()
    return render_template('products.html', products=products_list, search=search)

@app.route('/billing/products/add', methods=['POST'])
@login_required
def add_product():
    if current_user.role != 'admin':
        flash('Admin only.', 'danger')
        return redirect('/billing/products')
    price_in = float(request.form.get('price', 0))
    gst_rate = float(request.form.get('gst_rate', 18))
    price_type = request.form.get('price_type', 'without_gst')
    if price_type == 'with_gst':
        price = round(price_in * 100 / (100 + gst_rate), 2)
        price_with_gst = price_in
    else:
        price = price_in
        price_with_gst = round(price_in * (100 + gst_rate) / 100, 2)
    p = Product(
        name=request.form.get('name', '').strip(),
        description=request.form.get('description', '').strip(),
        price=price, price_with_gst=price_with_gst,
        stock=int(request.form.get('stock', 0)),
        gst_rate=gst_rate,
        hsn_code=request.form.get('hsn_code', '').strip(),
        brand=request.form.get('brand', '').strip(),
        category=request.form.get('category', '').strip(),
        unit=request.form.get('unit', 'Pcs').strip()
    )
    db.session.add(p)
    db.session.commit()
    flash('Product added!', 'success')
    return redirect('/billing/products')

@app.route('/billing/products/edit/<int:id>', methods=['POST'])
@login_required
def edit_product(id):
    if current_user.role != 'admin':
        flash('Admin only.', 'danger')
        return redirect('/billing/products')
    p = Product.query.get_or_404(id)
    p.name = request.form.get('name', p.name)
    p.description = request.form.get('description', p.description)
    price_in = float(request.form.get('price', p.price))
    gst_rate = float(request.form.get('gst_rate', p.gst_rate))
    price_type = request.form.get('price_type', 'without_gst')
    if price_type == 'with_gst':
        p.price = round(price_in * 100 / (100 + gst_rate), 2)
        p.price_with_gst = price_in
    else:
        p.price = price_in
        p.price_with_gst = round(price_in * (100 + gst_rate) / 100, 2)
    p.stock = int(request.form.get('stock', p.stock))
    p.gst_rate = gst_rate
    p.hsn_code = request.form.get('hsn_code', p.hsn_code)
    p.brand = request.form.get('brand', p.brand)
    p.category = request.form.get('category', p.category)
    p.unit = request.form.get('unit', p.unit)
    db.session.commit()
    flash('Product updated!', 'success')
    return redirect('/billing/products')

@app.route('/billing/products/delete/<int:id>', methods=['POST'])
@login_required
def delete_product(id):
    if current_user.role != 'admin':
        flash('Admin only.', 'danger')
        return redirect('/billing/products')
    p = Product.query.get_or_404(id)
    p.is_active = False
    db.session.commit()
    flash('Product deleted!', 'success')
    return redirect('/billing/products')

@app.route('/billing/products/delete-all', methods=['POST'])
@login_required
def delete_all_products():
    if current_user.role != 'admin':
        flash('Admin only.', 'danger')
        return redirect('/billing/products')
    Product.query.filter_by(is_active=True).update({'is_active': False})
    db.session.commit()
    flash('All products deleted!', 'success')
    return redirect('/billing/products')

@app.route('/billing/products/upload-excel', methods=['POST'])
@login_required
def upload_products_excel():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Admin only'}), 403
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file'}), 400
    f = request.files['file']
    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Excel file required (.xlsx or .xls)'}), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f)
        ws = wb.active
        added = 0
        errors = []
        default_price_type = request.form.get('price_type', 'without_gst')
        headers = [cell.value.strip().lower().replace(' ', '_') if cell.value else '' for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            data = dict(zip(headers, [str(c) if c is not None else '' for c in row]))
            try:
                gst = float(data.get('gst_rate', 18))
                price_in = float(data.get('price', 0))
                price_type = data.get('price_type', default_price_type)
                if price_type == 'with_gst':
                    price = round(price_in * 100 / (100 + gst), 2)
                    price_with_gst = price_in
                else:
                    price = price_in
                    price_with_gst = round(price_in * (100 + gst) / 100, 2)
                p = Product(
                    name=data.get('name', ''),
                    description=data.get('description', ''),
                    price=price, price_with_gst=price_with_gst,
                    stock=int(float(data.get('stock', 0))),
                    gst_rate=gst,
                    hsn_code=data.get('hsn_code', ''),
                    brand=data.get('brand', ''),
                    category=data.get('category', ''),
                    unit=data.get('unit', 'Pcs')
                )
                db.session.add(p)
                added += 1
            except Exception as e:
                errors.append(f'Row {added + 2}: {e}')
        db.session.commit()
        msg = f'{added} products added'
        if errors:
            msg += f', {len(errors)} errors: {"; ".join(errors[:3])}'
        flash(msg, 'success' if not errors else 'warning')
    except Exception as e:
        flash(f'Error reading Excel: {e}', 'danger')
    return redirect('/billing/products')

@app.route('/billing/billing')
@login_required
def billing():
    customers_list = Customer.query.all()
    products_list = Product.query.filter_by(is_active=True).all()
    invoices_list = Invoice.query.order_by(Invoice.date_created.desc()).limit(20).all()
    return render_template('billing.html', customers=customers_list, products=products_list, invoices=invoices_list)

@app.route('/billing/billing/create', methods=['POST'])
@login_required
def create_invoice():
    data = request.get_json()
    cust_name = data.get('customer_name', '').strip()
    cust_phone = data.get('customer_phone', '').strip()
    cust_email = data.get('customer_email', '')
    cust_gst = data.get('customer_gst', '')
    cust_state = data.get('customer_state', '')
    items = data.get('items', [])
    discount = float(data.get('discount', 0))
    is_igst = data.get('is_igst', False)
    reverse_charge = data.get('reverse_charge', False)
    due_date = data.get('due_date')
    place_of_supply = cust_state

    customer = None
    if cust_phone:
        customer = Customer.query.filter_by(phone=cust_phone).first()
    if not customer:
        customer = Customer(name=cust_name, phone=cust_phone, email=cust_email, gst_number=cust_gst, state=cust_state)
        db.session.add(customer)
        db.session.flush()

    inv_count = Invoice.query.count() + 1
    inv_number = f'INV-{inv_count:06d}'

    subtotal = 0
    total_cgst = 0
    total_sgst = 0
    total_igst = 0
    invoice_items = []

    for item_data in items:
        product = Product.query.get(item_data['product_id'])
        if not product:
            continue
        qty = int(item_data.get('quantity', 1))
        unit_price = float(item_data.get('unit_price', product.price))
        item_discount = float(item_data.get('discount', 0))
        gst_rate = float(item_data.get('gst_rate', product.gst_rate))
        taxable = unit_price * qty - item_discount
        gst_amount = taxable * gst_rate / 100
        total = taxable + gst_amount

        if is_igst:
            igst = gst_amount
            cgst_val = 0
            sgst_val = 0
        else:
            igst = 0
            cgst_val = gst_amount / 2
            sgst_val = gst_amount / 2

        inv_item = InvoiceItem(
            product_id=product.id, quantity=qty, unit_price=unit_price,
            discount=item_discount, taxable_value=taxable, gst_rate=gst_rate,
            gst_amount=gst_amount, igst=igst, total=total,
            star_rating=item_data.get('star_rating', ''),
            model_no=item_data.get('model_no', ''),
            series_no=item_data.get('series_no', ''),
            item_category=item_data.get('item_category', product.category or ''),
            ton_size=item_data.get('ton_size', '')
        )
        invoice_items.append(inv_item)
        subtotal += taxable
        total_cgst += cgst_val
        total_sgst += sgst_val
        total_igst += igst
        product.stock -= qty

    total_tax = total_cgst + total_sgst + total_igst
    grand = subtotal + total_tax - discount
    round_off = round(grand) - grand
    grand_total = round(grand)

    invoice = Invoice(
        invoice_number=inv_number, customer_id=customer.id, user_id=current_user.id,
        subtotal=subtotal, cgst=total_cgst, sgst=total_sgst, igst=total_igst,
        total_tax=total_tax, total_amount=subtotal, discount=discount,
        round_off=round_off, grand_total=grand_total, status='unpaid',
        place_of_supply=place_of_supply, reverse_charge=reverse_charge
    )
    db.session.add(invoice)
    db.session.flush()
    for ii in invoice_items:
        ii.invoice_id = invoice.id
        db.session.add(ii)
    db.session.commit()

    return jsonify({'success': True, 'invoice_id': invoice.id, 'invoice_number': inv_number})

@app.route('/billing/invoice/<int:id>/delete', methods=['POST'])
@login_required
def delete_invoice(id):
    if current_user.role != 'admin':
        flash('Admin only.', 'danger')
        return redirect('/billing/billing')
    inv = Invoice.query.get_or_404(id)
    for item in inv.items:
        prod = Product.query.get(item.product_id)
        if prod:
            prod.stock += item.quantity
    Payment.query.filter_by(invoice_id=inv.id).delete()
    InvoiceItem.query.filter_by(invoice_id=inv.id).delete()
    db.session.delete(inv)
    db.session.commit()
    flash('Invoice deleted!', 'success')
    return redirect('/billing/billing')

@app.route('/billing/invoice/<int:id>/pdf')
@login_required
def download_invoice_pdf(id):
    from fpdf import FPDF
    invoice = Invoice.query.get_or_404(id)
    company = CompanySettings.query.first()
    if not company:
        company = CompanySettings()

    paid_amount = sum(p.amount for p in invoice.payments)
    is_paid = (invoice.grand_total - paid_amount) <= 0
    status_text = 'PAID' if invoice.status == 'paid' else ('PARTIAL' if invoice.status == 'partial' else 'UNPAID')

    def _w(n):
        a = ['','One','Two','Three','Four','Five','Six','Seven','Eight','Nine','Ten',
             'Eleven','Twelve','Thirteen','Fourteen','Fifteen','Sixteen','Seventeen','Eighteen','Nineteen']
        b = ['','','Twenty','Thirty','Forty','Fifty','Sixty','Seventy','Eighty','Ninety']
        if n < 20: return a[n]
        return (b[n // 10] + ' ' + a[n % 10]).strip()

    def amount_words(num):
        n = int(num)
        if n == 0: return 'Zero'
        parts = []
        if n >= 10000000: parts.append(_w(n // 10000000) + ' Crore'); n %= 10000000
        if n >= 100000: parts.append(_w(n // 100000) + ' Lakh'); n %= 100000
        if n >= 1000: parts.append(_w(n // 1000) + ' Thousand'); n %= 1000
        if n >= 100: parts.append(_w(n // 100) + ' Hundred'); n %= 100
        if n > 0: parts.append(_w(n))
        return ' '.join(parts)
    dec = round((invoice.grand_total - int(invoice.grand_total)) * 100)
    words = amount_words(int(invoice.grand_total))
    if dec > 0: words += ' and ' + amount_words(dec) + ' Paise'
    words += ' Only'

    class PrintInvoicePDF(FPDF):
        def footer(self):
            self.set_y(-18)
            self.set_draw_color(150, 150, 150)
            self.line(10, self.get_y(), 200, self.get_y())
            self.ln(3)
            self.set_font('Helvetica', 'B', 8)
            self.set_text_color(50, 50, 50)
            self.cell(90, 5, 'Authorized Signatory', align='C')
            self.cell(90, 5, company.company_name or '', align='C')

    pdf = PrintInvoicePDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=22)

    # ===== WATERMARK =====
    pdf.set_font('Helvetica', 'B', 56)
    pdf.set_text_color(230, 230, 245)
    pdf.set_xy(20, 80)
    pdf.cell(170, 30, 'INDIAN A/C', align='C')

    pdf.set_xy(10, 10)

    # ===== HEADER =====
    # Left side: logo + company info
    y_header = pdf.get_y()
    if company.logo_filename:
        logo_path = os.path.join(BILLING_DIR, 'static', 'uploads', company.logo_filename)
        if os.path.exists(logo_path):
            try:
                pdf.image(logo_path, 10, 10, 25)
            except Exception:
                pass

    pdf.set_xy(10, 12)
    pdf.set_font('Helvetica', 'B', 16)
    pdf.set_text_color(255, 140, 0)
    pdf.cell(100, 8, company.company_name or 'Indian A/C Sales and Services', new_x='LMARGIN', new_y='NEXT')

    if company.tagline:
        pdf.set_x(10)
        pdf.set_font('Helvetica', 'B', 7.5)
        pdf.set_text_color(136, 136, 136)
        pdf.cell(100, 4, company.tagline, new_x='LMARGIN', new_y='NEXT')

    pdf.set_x(10)
    pdf.set_font('Helvetica', '', 7.5)
    pdf.set_text_color(85, 85, 85)
    line = ''
    if company.address: line += company.address
    if company.city: line += ', ' + company.city
    if company.state: line += ', ' + company.state
    if company.pincode: line += ' - ' + company.pincode
    if line: pdf.cell(100, 4, line.strip(', '), new_x='LMARGIN', new_y='NEXT')

    detail_line = ''
    if company.phone: detail_line += 'Phone: ' + company.phone
    if company.email: detail_line += ' | Email: ' + company.email
    if detail_line:
        pdf.set_x(10)
        pdf.cell(100, 4, detail_line.strip(' | '), new_x='LMARGIN', new_y='NEXT')
    if company.website:
        pdf.set_x(10)
        pdf.cell(100, 4, company.website, new_x='LMARGIN', new_y='NEXT')
    gst_pan = ''
    if company.gstin: gst_pan += 'GSTIN: ' + company.gstin
    if company.pan: gst_pan += ' | PAN: ' + company.pan
    if gst_pan:
        pdf.set_x(10)
        pdf.set_font('Helvetica', 'B', 7.5)
        pdf.cell(100, 4, gst_pan.strip(' | '), new_x='LMARGIN', new_y='NEXT')

    # Right side: TAX INVOICE + invoice info
    pdf.set_xy(130, 10)
    pdf.set_font('Helvetica', 'B', 16)
    pdf.set_text_color(0, 0, 128)
    pdf.cell(70, 8, 'TAX INVOICE', align='R', new_x='LMARGIN', new_y='NEXT')

    pdf.set_xy(130, 19)
    pdf.set_font('Helvetica', '', 8.5)
    pdf.set_text_color(50, 50, 50)
    pdf.cell(70, 4.5, f'Invoice #: {invoice.invoice_number}', align='R', new_x='LMARGIN', new_y='NEXT')
    pdf.set_xy(130, pdf.get_y())
    pdf.cell(70, 4.5, f'Date: {invoice.date_created.strftime("%d-%m-%Y %H:%M")}', align='R', new_x='LMARGIN', new_y='NEXT')
    pdf.set_xy(130, pdf.get_y())
    pdf.cell(70, 4.5, f'Status: {status_text}', align='R', new_x='LMARGIN', new_y='NEXT')
    if invoice.user:
        pdf.set_xy(130, pdf.get_y())
        pdf.cell(70, 4.5, f'Prepared by: {invoice.user.full_name}', align='R', new_x='LMARGIN', new_y='NEXT')

    # Navy blue bottom border
    header_end = max(pdf.get_y(), 42) + 2
    pdf.set_draw_color(0, 0, 128)
    pdf.set_line_width(0.8)
    pdf.line(10, header_end, 200, header_end)
    pdf.set_line_width(0.2)
    pdf.set_y(header_end + 3)

    # ===== BILL TO BOX =====
    pdf.set_fill_color(248, 249, 250)
    pdf.set_draw_color(233, 236, 239)
    box_y = pdf.get_y()
    pdf.rect(10, box_y, 190, 28, style='DF')
    pdf.set_xy(13, box_y + 2)
    pdf.set_font('Helvetica', 'B', 8.5)
    pdf.set_text_color(0, 0, 128)
    pdf.cell(40, 4, 'Bill To:', new_x='LMARGIN', new_y='NEXT')
    pdf.set_x(13)
    pdf.set_font('Helvetica', 'B', 8.5)
    pdf.set_text_color(50, 50, 50)
    pdf.cell(90, 4, invoice.customer.name, new_x='LMARGIN', new_y='NEXT')
    pdf.set_x(13)
    pdf.set_font('Helvetica', '', 7.5)
    pdf.set_text_color(85, 85, 85)
    if invoice.customer.address:
        pdf.cell(90, 4, invoice.customer.address, new_x='LMARGIN', new_y='NEXT')
        pdf.set_x(13)
    if invoice.customer.phone:
        pdf.cell(90, 4, f'Phone: {invoice.customer.phone}', new_x='LMARGIN', new_y='NEXT')
        pdf.set_x(13)
    if invoice.customer.email:
        pdf.cell(90, 4, f'Email: {invoice.customer.email}', new_x='LMARGIN', new_y='NEXT')
        pdf.set_x(13)
    if invoice.customer.gst_number:
        pdf.set_font('Helvetica', 'B', 7.5)
        pdf.cell(90, 4, f'GSTIN: {invoice.customer.gst_number}', new_x='LMARGIN', new_y='NEXT')
        pdf.set_x(13)
    if invoice.place_of_supply:
        pdf.set_font('Helvetica', 'B', 7.5)
        pdf.cell(90, 4, f'Place of Supply: {invoice.place_of_supply}', new_x='LMARGIN', new_y='NEXT')
        pdf.set_x(13)
    pdf.set_y(box_y + 30)

    # ===== ITEMS TABLE =====
    pdf.set_font('Helvetica', 'B', 7)
    pdf.set_fill_color(0, 0, 128)
    pdf.set_text_color(255, 255, 255)
    col_w = [8, 38, 16, 12, 10, 18, 14, 18, 12, 18, 24]
    headers = ['#', 'Name', 'HSN', 'Unit', 'Qty', 'Rate', 'Disc', 'Taxable', 'GST%', 'GST Amt', 'Total']
    for i, h in enumerate(headers):
        align = 'C' if i in [0, 3, 4, 8] else ('R' if i >= 5 else 'L')
        pdf.cell(col_w[i], 6, h, border=1, align=align, fill=True)
    pdf.ln()

    pdf.set_font('Helvetica', '', 7)
    pdf.set_text_color(50, 50, 50)
    for idx, item in enumerate(invoice.items, 1):
        if idx % 2 == 0:
            pdf.set_fill_color(248, 249, 250)
            fill = True
        else:
            fill = False
        taxable = item.taxable_value if item.taxable_value else (item.unit_price * item.quantity)
        unit = getattr(item, 'unit', None) or getattr(item.product, 'unit', 'NOS') or 'NOS'
        desc = []
        if item.product.description: desc.append(item.product.description[:20])
        if item.product.brand: desc.append('Brand:' + item.product.brand)
        if item.product.category: desc.append(item.product.category)
        prod_text = (item.product.name or '')[:20]
        if desc: prod_text += '\n' + ' '.join(desc)[:30]
        row = [
            str(idx),
            prod_text,
            item.product.hsn_code or '-',
            unit,
            str(item.quantity),
            f'{item.unit_price:,.2f}',
            f'{item.discount:,.2f}' if item.discount else '-',
            f'{taxable:,.2f}',
            f'{item.gst_rate}%',
            f'{item.gst_amount:,.2f}',
            f'{item.total:,.2f}'
        ]
        # Save x position, use multi_cell for product column to allow newline
        x0 = pdf.get_x()
        y0 = pdf.get_y()
        row_h = 8 if desc else 5.5
        for i, val in enumerate(row):
            pdf.set_xy(x0 + sum(col_w[:i]), y0)
            align = 'C' if i in [0, 3, 4, 8] else ('R' if i >= 5 else 'L')
            if i == 1:
                pdf.multi_cell(col_w[i], row_h / 2, val, border=1, align=align, fill=fill)
            else:
                pdf.cell(col_w[i], row_h, val, border=1, align=align, fill=fill)
        if desc:
            pdf.set_y(max(pdf.get_y(), y0 + row_h))

    pdf.ln(3)

    # ===== SUMMARY SECTION =====
    summary_y = pdf.get_y()

    # Left side: Payment History
    pdf.set_x(10)
    if invoice.payments:
        pdf.set_font('Helvetica', 'B', 7.5)
        pdf.set_fill_color(0, 0, 128)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(95, 5.5, '  Payment History', border=1, fill=True)
        pdf.ln()
        pdf.set_font('Helvetica', 'B', 6.5)
        pdf.set_fill_color(233, 236, 239)
        pdf.set_text_color(50, 50, 50)
        pdf.cell(28, 5, '  Date', border=1, fill=True)
        pdf.cell(25, 5, '  Method', border=1, fill=True)
        pdf.cell(22, 5, '  Reference', border=1, fill=True)
        pdf.cell(20, 5, 'Amount', border=1, fill=True, align='R')
        pdf.ln()
        pdf.set_font('Helvetica', '', 6.5)
        for p in invoice.payments:
            pdf.cell(28, 4.5, '  ' + p.payment_date.strftime('%d-%m-%Y'), border=1)
            pdf.cell(25, 4.5, '  ' + (p.payment_method or '').title(), border=1)
            pdf.cell(22, 4.5, '  ' + (p.reference or '-'), border=1)
            pdf.cell(20, 4.5, f'{p.amount:,.2f}', border=1, align='R')
            pdf.ln()
    else:
        pdf.set_font('Helvetica', 'I', 7.5)
        pdf.set_text_color(136, 136, 136)
        pdf.cell(95, 5, 'No payments recorded.', align='C')
        pdf.ln()

    # Right side: Totals
    pdf.set_xy(145, summary_y)
    pdf.set_font('Helvetica', '', 8.5)
    pdf.set_text_color(50, 50, 50)
    pdf.cell(30, 5.5, 'Subtotal:')
    pdf.cell(25, 5.5, f'{invoice.subtotal:,.2f}', align='R', new_x='LMARGIN', new_y='NEXT')
    if invoice.igst and invoice.igst > 0:
        pdf.set_x(145)
        pdf.cell(30, 5.5, 'IGST:')
        pdf.cell(25, 5.5, f'{invoice.igst:,.2f}', align='R', new_x='LMARGIN', new_y='NEXT')
    else:
        if invoice.cgst:
            pdf.set_x(145)
            pdf.cell(30, 5.5, 'CGST:')
            pdf.cell(25, 5.5, f'{invoice.cgst:,.2f}', align='R', new_x='LMARGIN', new_y='NEXT')
        if invoice.sgst:
            pdf.set_x(145)
            pdf.cell(30, 5.5, 'SGST:')
            pdf.cell(25, 5.5, f'{invoice.sgst:,.2f}', align='R', new_x='LMARGIN', new_y='NEXT')
    if invoice.discount and invoice.discount > 0:
        pdf.set_x(145)
        pdf.cell(30, 5.5, 'Discount:')
        pdf.cell(25, 5.5, f'-{invoice.discount:,.2f}', align='R', new_x='LMARGIN', new_y='NEXT')

    # Grand Total with navy top border
    pdf.set_x(145)
    pdf.set_draw_color(0, 0, 128)
    pdf.set_line_width(0.4)
    pdf.line(145, pdf.get_y(), 195, pdf.get_y())
    pdf.set_line_width(0.2)
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_text_color(0, 0, 128)
    pdf.cell(30, 7, 'Grand Total:')
    pdf.cell(25, 7, f'{invoice.grand_total:,.2f}', align='R', new_x='LMARGIN', new_y='NEXT')

    pdf.set_x(145)
    pdf.set_font('Helvetica', '', 8.5)
    pdf.set_text_color(40, 167, 69)
    pdf.cell(30, 5.5, 'Paid:')
    pdf.cell(25, 5.5, f'{paid_amount:,.2f}', align='R', new_x='LMARGIN', new_y='NEXT')

    due = invoice.grand_total - paid_amount
    if due > 0:
        pdf.set_x(145)
        pdf.set_font('Helvetica', 'B', 8.5)
        pdf.set_text_color(220, 53, 69)
        pdf.cell(30, 5.5, 'Due:')
        pdf.cell(25, 5.5, f'{due:,.2f}', align='R', new_x='LMARGIN', new_y='NEXT')

    pdf.set_y(max(pdf.get_y(), summary_y) + 12)

    # ===== AMOUNT IN WORDS =====
    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_text_color(0, 0, 128)
    pdf.cell(25, 5, 'Amount in Words:')
    pdf.set_font('Helvetica', 'I', 8)
    pdf.set_text_color(50, 50, 50)
    pdf.cell(0, 5, words, new_x='LMARGIN', new_y='NEXT')
    pdf.ln(3)

    # ===== TERMS =====
    pdf.set_fill_color(248, 249, 250)
    pdf.set_draw_color(233, 236, 239)
    terms_y = pdf.get_y()
    pdf.rect(10, terms_y, 190, 14, style='DF')
    pdf.set_xy(13, terms_y + 2)
    pdf.set_font('Helvetica', 'B', 7.5)
    pdf.set_text_color(50, 50, 50)
    pdf.cell(0, 4, 'Terms & Conditions:', new_x='LMARGIN', new_y='NEXT')
    pdf.set_x(13)
    pdf.set_font('Helvetica', '', 7)
    pdf.cell(0, 4, 'Bill is mandatory to claim warranty. Please retain this invoice for future reference.', new_x='LMARGIN', new_y='NEXT')
    pdf.set_x(13)
    pdf.cell(0, 4, 'Any disputes shall be subject to local jurisdiction only.', new_x='LMARGIN', new_y='NEXT')
    pdf.set_y(terms_y + 16)

    # ===== PAID/DUE STAMP =====
    stamp_x = 168
    stamp_y = pdf.get_y() + 1
    r = 14
    color = (40, 167, 69) if is_paid else (220, 53, 69)
    pdf.set_draw_color(*color)
    pdf.set_line_width(0.6)
    pdf.circle(stamp_x, stamp_y, r, style='D')
    pdf.set_line_width(0.3)
    pdf.circle(stamp_x, stamp_y, r - 2, style='D')
    stamp_text = 'RECEIVED' if is_paid else 'DUE'
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_text_color(*color)
    pdf.set_xy(stamp_x - r + 2, stamp_y - 4)
    pdf.cell(r * 2 - 4, 5, stamp_text, align='C')
    pdf.set_font('Helvetica', '', 6)
    pdf.set_xy(stamp_x - r + 2, stamp_y + 1)
    pdf.cell(r * 2 - 4, 4, invoice.date_created.strftime('%d-%m-%Y'), align='C')

    pdf.set_y(stamp_y + r + 5)

    # ===== FOOTER =====
    pdf.set_draw_color(233, 236, 239)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(2)
    pdf.set_font('Helvetica', '', 7.5)
    pdf.set_text_color(136, 136, 136)
    pdf.cell(0, 4, f'Thank you for your business with {company.company_name or "Indian A/C Sales and Services"}!', align='C', new_x='LMARGIN', new_y='NEXT')
    pdf.cell(0, 4, 'This is a computer-generated invoice.', align='C')

    buf = BytesIO()
    pdf_bytes = pdf.output()
    if isinstance(pdf_bytes, str):
        pdf_bytes = pdf_bytes.encode('latin-1')
    buf.write(pdf_bytes)
    buf.seek(0)

    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f'{invoice.invoice_number}.pdf')

@app.route('/billing/payments')
@login_required
def payments():
    search = request.args.get('search', '').strip()
    q = Invoice.query
    if search:
        q = q.filter(Invoice.invoice_number.ilike(f'%{search}%'))
    invoices = q.order_by(Invoice.date_created.desc()).all()
    return render_template('payments.html', invoices=invoices, search=search)

@app.route('/billing/payments/add', methods=['POST'])
@login_required
def add_payment():
    inv_id = int(request.form.get('invoice_id', 0))
    amount = float(request.form.get('amount', 0))
    method = request.form.get('payment_method', 'cash')
    reference = request.form.get('reference', '')
    inv = Invoice.query.get_or_404(inv_id)
    p = Payment(invoice_id=inv_id, amount=amount, payment_method=method, reference=reference)
    db.session.add(p)
    total_paid = sum(x.amount for x in inv.payments) + amount
    due = inv.grand_total - total_paid
    if total_paid >= inv.grand_total:
        inv.status = 'paid'
    else:
        inv.status = 'partial'
    db.session.commit()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'paid': total_paid, 'due': max(0, due), 'status': inv.status, 'payment_id': p.id})
    flash('Payment recorded!', 'success')
    return redirect('/billing/payments')

@app.route('/billing/payments/delete/<int:id>', methods=['POST'])
@login_required
def delete_payment(id):
    if current_user.role != 'admin':
        flash('Admin only.', 'danger')
        return redirect('/billing/payments')
    p = Payment.query.get_or_404(id)
    inv = Invoice.query.get(p.invoice_id)
    db.session.delete(p)
    if inv:
        total_paid = sum(x.amount for x in inv.payments if x.id != id)
        if total_paid >= inv.grand_total:
            inv.status = 'paid'
        elif total_paid > 0:
            inv.status = 'partial'
        else:
            inv.status = 'unpaid'
    db.session.commit()
    flash('Payment deleted!', 'success')
    return redirect('/billing/payments')

@app.route('/billing/attendance')
@login_required
def attendance():
    selected_date = request.args.get('date', date.today().isoformat())
    att_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
    employees_list = BEmployee.query.all()
    attendance_records = {}
    for att in BAttendance.query.filter_by(date=att_date).all():
        attendance_records[att.employee_id] = att
    return render_template('attendance.html', employees=employees_list, selected_date=selected_date, attendance_records=attendance_records)

@app.route('/billing/attendance/mark', methods=['POST'])
@login_required
def mark_attendance():
    if current_user.role != 'admin':
        flash('Admin only.', 'danger')
        return redirect('/billing/attendance')
    emp_id = int(request.form.get('employee_id', 0))
    att_date = request.form.get('date', date.today().isoformat())
    check_in = request.form.get('check_in', '')
    check_out = request.form.get('check_out', '')
    status = request.form.get('status', 'present')

    att = BAttendance.query.filter_by(employee_id=emp_id, date=datetime.strptime(att_date, '%Y-%m-%d').date()).first()
    if not att:
        att = BAttendance(employee_id=emp_id, date=datetime.strptime(att_date, '%Y-%m-%d').date())
        db.session.add(att)
    if check_in:
        att.check_in = datetime.strptime(f'{att_date} {check_in}', '%Y-%m-%d %H:%M')
    if check_out:
        att.check_out = datetime.strptime(f'{att_date} {check_out}', '%Y-%m-%d %H:%M')
    att.status = status
    db.session.commit()
    flash('Attendance marked!', 'success')
    return redirect('/billing/attendance')

@app.route('/billing/reports')
@login_required
def reports():
    now = datetime.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    year_start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)

    monthly_sales = db.session.query(db.func.coalesce(db.func.sum(Invoice.grand_total), 0)).filter(Invoice.date_created >= month_start).scalar()
    yearly_sales = db.session.query(db.func.coalesce(db.func.sum(Invoice.grand_total), 0)).filter(Invoice.date_created >= year_start).scalar()
    total_tax_collected = db.session.query(db.func.coalesce(db.func.sum(Invoice.total_tax), 0)).scalar()

    monthly_data = []
    for i in range(11, -1, -1):
        m = (now.month - i - 1) % 12 + 1
        y = now.year if now.month - i > 0 else now.year - 1
        ms = datetime(y, m, 1)
        me = datetime(y + (m // 12), (m % 12) + 1, 1)
        rev = db.session.query(db.func.coalesce(db.func.sum(Invoice.grand_total), 0)).filter(Invoice.date_created >= ms, Invoice.date_created < me).scalar()
        monthly_data.append({'month': ms.strftime('%b %Y'), 'revenue': float(rev)})

    payment_methods = [[row[0], row[1]] for row in db.session.query(Payment.payment_method, db.func.count(Payment.id)).group_by(Payment.payment_method).all()]

    product_sales = db.session.query(Product.name, db.func.sum(InvoiceItem.quantity)).join(InvoiceItem, InvoiceItem.product_id == Product.id).group_by(Product.name).order_by(db.func.sum(InvoiceItem.quantity).desc()).limit(10).all()
    top_products = [(name, int(qty)) for name, qty in product_sales]

    invoices = Invoice.query.order_by(Invoice.date_created.desc()).all()
    sales_history = []
    for inv in invoices:
        paid = sum(p.amount for p in inv.payments)
        status = 'paid' if paid >= inv.grand_total else 'partial' if paid > 0 else 'unpaid'
        sales_history.append({
            'id': inv.id,
            'invoice_number': inv.invoice_number,
            'date_created': inv.date_created,
            'customer_name': inv.customer.name if inv.customer else 'Unknown',
            'emp_username': inv.user.username if inv.user else '-',
            'emp_name': inv.user.full_name if inv.user else '-',
            'grand_total': inv.grand_total,
            'status': status
        })

    return render_template('reports.html', monthly_sales=monthly_sales, yearly_sales=yearly_sales,
                           total_tax_collected=total_tax_collected, monthly_data=monthly_data,
                           payment_methods=payment_methods, top_products=top_products, sales_history=sales_history)

@app.route('/billing/reports/download')
@login_required
def download_report():
    from fpdf import FPDF
    company = CompanySettings.query.first()
    if not company:
        company = CompanySettings()

    today = date.today()
    month_start = today.replace(day=1)
    year_start = today.replace(month=1, day=1)
    monthly_sales = db.session.query(db.func.coalesce(db.func.sum(Invoice.grand_total), 0)).filter(Invoice.date_created >= datetime.combine(month_start, datetime.min.time())).scalar() or 0
    yearly_sales = db.session.query(db.func.coalesce(db.func.sum(Invoice.grand_total), 0)).filter(Invoice.date_created >= datetime.combine(year_start, datetime.min.time())).scalar() or 0
    total_tax = db.session.query(db.func.coalesce(db.func.sum(Invoice.total_tax), 0)).filter(Invoice.date_created >= datetime.combine(year_start, datetime.min.time())).scalar() or 0
    total_invoices = Invoice.query.filter(Invoice.date_created >= datetime.combine(year_start, datetime.min.time())).count()

    all_invoices = Invoice.query.filter(Invoice.date_created >= datetime.combine(year_start, datetime.min.time())).order_by(Invoice.date_created.desc()).all()

    product_sales = db.session.query(
        Product.name, db.func.sum(InvoiceItem.quantity)
    ).join(InvoiceItem, InvoiceItem.product_id == Product.id
    ).join(Invoice, Invoice.id == InvoiceItem.invoice_id
    ).filter(Invoice.date_created >= datetime.combine(year_start, datetime.min.time())
    ).group_by(Product.name).order_by(db.func.sum(InvoiceItem.quantity).desc()).limit(10).all()

    class ReportPDF(FPDF):
        def header(self):
            if company.logo_filename:
                logo_path = os.path.join(BILLING_DIR, 'static', 'uploads', company.logo_filename)
                if os.path.exists(logo_path):
                    try:
                        self.image(logo_path, 10, 8, 25)
                    except Exception:
                        pass
            self.set_font('Helvetica', 'B', 14)
            self.set_text_color(255, 140, 0)
            self.cell(0, 8, company.company_name or 'Indian A/C Sales and Services', new_x='LMARGIN', new_y='NEXT')
            self.set_font('Helvetica', '', 7)
            self.set_text_color(100, 100, 100)
            addr = company.address or ''
            if company.city: addr += ', ' + company.city
            if company.state: addr += ', ' + company.state
            if company.pincode: addr += ' - ' + company.pincode
            if addr: self.cell(0, 4, addr.strip(', '), new_x='LMARGIN', new_y='NEXT')
            if company.gstin: self.cell(0, 4, 'GSTIN: ' + company.gstin, new_x='LMARGIN', new_y='NEXT')
            self.ln(2)
            self.set_draw_color(200, 200, 200)
            self.line(10, self.get_y(), 200, self.get_y())
            self.ln(4)

        def footer(self):
            self.set_y(-15)
            self.set_font('Helvetica', 'I', 7)
            self.set_text_color(150, 150, 150)
            self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', align='C')

    pdf = ReportPDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=20)

    # Title
    pdf.set_font('Helvetica', 'B', 16)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 10, 'SALES REPORT', new_x='LMARGIN', new_y='NEXT', align='C')
    pdf.set_font('Helvetica', '', 8)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 5, f'Generated: {today.strftime("%d-%m-%Y")} | Period: {year_start.strftime("%d %b %Y")} - {today.strftime("%d %b %Y")}', new_x='LMARGIN', new_y='NEXT', align='C')
    pdf.ln(6)

    # Summary cards
    pdf.set_font('Helvetica', 'B', 9)
    card_w = 60
    pdf.set_fill_color(230, 240, 255)
    pdf.cell(card_w, 18, '', border=1, fill=True, align='C')
    pdf.set_xy(10, pdf.get_y() - 16)
    pdf.set_text_color(0, 0, 128)
    pdf.cell(card_w, 8, f'Rs. {monthly_sales:,.2f}', align='C')
    pdf.set_xy(10, pdf.get_y() + 8)
    pdf.set_text_color(80, 80, 80)
    pdf.set_font('Helvetica', '', 7)
    pdf.cell(card_w, 5, 'Monthly Sales', align='C')

    pdf.set_xy(75, 18)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_fill_color(230, 255, 230)
    pdf.cell(card_w, 18, '', border=1, fill=True, align='C')
    pdf.set_xy(75, 20)
    pdf.set_text_color(0, 128, 0)
    pdf.cell(card_w, 8, f'Rs. {yearly_sales:,.2f}', align='C')
    pdf.set_xy(75, 28)
    pdf.set_text_color(80, 80, 80)
    pdf.set_font('Helvetica', '', 7)
    pdf.cell(card_w, 5, 'Yearly Sales', align='C')

    pdf.set_xy(140, 18)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_fill_color(255, 245, 220)
    pdf.cell(card_w, 18, '', border=1, fill=True, align='C')
    pdf.set_xy(140, 20)
    pdf.set_text_color(180, 120, 0)
    pdf.cell(card_w, 8, f'Rs. {total_tax:,.2f}', align='C')
    pdf.set_xy(140, 28)
    pdf.set_text_color(80, 80, 80)
    pdf.set_font('Helvetica', '', 7)
    pdf.cell(card_w, 5, 'Total Tax Collected', align='C')

    pdf.set_y(42)
    pdf.ln(4)

    # Top Products
    if product_sales:
        pdf.set_font('Helvetica', 'B', 10)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 7, 'Top Selling Products', new_x='LMARGIN', new_y='NEXT')
        pdf.set_font('Helvetica', 'B', 7)
        pdf.set_fill_color(40, 40, 40)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(10, 6, '#', border=1, fill=True, align='C')
        pdf.cell(100, 6, 'Product', border=1, fill=True)
        pdf.cell(30, 6, 'Qty Sold', border=1, fill=True, align='R')
        pdf.ln()
        pdf.set_font('Helvetica', '', 7)
        pdf.set_text_color(0, 0, 0)
        for i, (name, qty) in enumerate(product_sales, 1):
            fill = i % 2 == 0
            if fill: pdf.set_fill_color(245, 245, 245)
            pdf.cell(10, 5, str(i), border=1, align='C', fill=fill)
            pdf.cell(100, 5, name[:45], border=1, fill=fill)
            pdf.cell(30, 5, str(int(qty)), border=1, align='R', fill=fill)
            pdf.ln()
        pdf.ln(4)

    # Sales History table
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 7, 'Sales History', new_x='LMARGIN', new_y='NEXT')

    pdf.set_font('Helvetica', 'B', 6)
    pdf.set_fill_color(40, 40, 40)
    pdf.set_text_color(255, 255, 255)
    cols = [8, 25, 22, 35, 30, 25, 25, 20]
    heads = ['#', 'Invoice', 'Date', 'Customer', 'Staff', 'Amount', 'Tax', 'Status']
    for i, h in enumerate(heads):
        pdf.cell(cols[i], 6, h, border=1, fill=True, align='C')
    pdf.ln()

    pdf.set_font('Helvetica', '', 6)
    pdf.set_text_color(0, 0, 0)
    for idx, inv in enumerate(all_invoices[:50], 1):
        fill = idx % 2 == 0
        if fill: pdf.set_fill_color(245, 245, 245)
        pdf.cell(cols[0], 5, str(idx), border=1, align='C', fill=fill)
        pdf.cell(cols[1], 5, inv.invoice_number, border=1, fill=fill)
        pdf.cell(cols[2], 5, inv.date_created.strftime('%d-%m-%y'), border=1, fill=fill)
        cust_name = inv.customer.name if inv.customer else '-'
        pdf.cell(cols[3], 5, cust_name[:18], border=1, fill=fill)
        emp_name = inv.user.full_name if inv.user else '-'
        pdf.cell(cols[4], 5, emp_name[:16], border=1, fill=fill)
        pdf.cell(cols[5], 5, f'{inv.grand_total:,.0f}', border=1, align='R', fill=fill)
        pdf.cell(cols[6], 5, f'{inv.total_tax:,.0f}', border=1, align='R', fill=fill)
        status_text = inv.status.upper()
        pdf.cell(cols[7], 5, status_text, border=1, align='C', fill=fill)
        pdf.ln()

    if not all_invoices:
        pdf.set_font('Helvetica', 'I', 8)
        pdf.set_text_color(150, 150, 150)
        pdf.cell(0, 10, 'No sales records found', new_x='LMARGIN', new_y='NEXT', align='C')

    # Signature
    pdf.ln(10)
    pdf.set_font('Helvetica', '', 8)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(90, 5, '', align='L')
    pdf.cell(90, 5, 'For ' + (company.company_name or ''), align='R', new_x='LMARGIN', new_y='NEXT')
    pdf.ln(15)
    pdf.cell(90, 5, '', align='L')
    pdf.cell(90, 5, 'Authorized Signatory', align='R', new_x='LMARGIN', new_y='NEXT')

    buf = BytesIO()
    pdf_bytes = pdf.output()
    if isinstance(pdf_bytes, str):
        pdf_bytes = pdf_bytes.encode('latin-1')
    buf.write(pdf_bytes)
    buf.seek(0)

    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f'Sales_Report_{today.strftime("%d%m%Y")}.pdf')

@app.route('/billing/employees')
@login_required
def employees():
    employees_list = BEmployee.query.all()
    return render_template('employees.html', employees=employees_list)

@app.route('/billing/employees/add', methods=['POST'])
@login_required
def add_employee():
    if current_user.role != 'admin':
        flash('Admin only.', 'danger')
        return redirect('/billing/employees')
    username = request.form.get('username', '').strip()
    password = request.form.get('password', 'staff123')
    full_name = request.form.get('full_name', '').strip()
    role = request.form.get('role', 'staff')
    user = BUser(username=username, password=generate_password_hash(password), full_name=full_name, role=role)
    db.session.add(user)
    db.session.flush()
    emp_count = BEmployee.query.count()
    emp = BEmployee(user_id=user.id, employee_id=f'EMP-{emp_count+1:05d}', designation=request.form.get('designation', ''), department=request.form.get('department', ''), salary=float(request.form.get('salary', 0)), phone=request.form.get('phone', ''), address=request.form.get('address', ''))
    db.session.add(emp)
    db.session.commit()
    flash('Employee added!', 'success')
    return redirect('/billing/employees')

@app.route('/billing/employees/edit/<int:id>', methods=['POST'])
@login_required
def edit_employee(id):
    if current_user.role != 'admin':
        flash('Admin only.', 'danger')
        return redirect('/billing/employees')
    emp = BEmployee.query.get_or_404(id)
    user = emp.user
    user.full_name = request.form.get('full_name', user.full_name)
    user.role = request.form.get('role', user.role)
    emp.designation = request.form.get('designation', emp.designation)
    emp.department = request.form.get('department', emp.department)
    emp.salary = float(request.form.get('salary', emp.salary))
    emp.phone = request.form.get('phone', emp.phone)
    emp.address = request.form.get('address', emp.address)
    db.session.commit()
    flash('Employee updated!', 'success')
    return redirect('/billing/employees')

@app.route('/billing/employees/delete/<int:id>', methods=['POST'])
@login_required
def delete_employee(id):
    if current_user.role != 'admin':
        flash('Admin only.', 'danger')
        return redirect('/billing/employees')
    emp = BEmployee.query.get_or_404(id)
    user = emp.user
    db.session.delete(emp)
    if user:
        BAttendance.query.filter_by(employee_id=emp.id).delete()
        db.session.delete(user)
    db.session.commit()
    flash('Employee deleted!', 'success')
    return redirect('/billing/employees')

@app.route('/billing/employees/reset-password/<int:id>', methods=['POST'])
@login_required
def reset_employee_password(id):
    if current_user.role != 'admin':
        flash('Admin only.', 'danger')
        return redirect('/billing/employees')
    emp = BEmployee.query.get_or_404(id)
    new_pass = request.form.get('new_password', '')
    if new_pass:
        emp.user.password = generate_password_hash(new_pass)
        db.session.commit()
        flash('Password reset!', 'success')
    return redirect('/billing/employees')

@app.route('/billing/settings', methods=['GET', 'POST'])
@login_required
def settings():
    if request.method == 'POST':
        current_user.full_name = request.form.get('full_name', current_user.full_name)
        new_pass = request.form.get('new_password', '')
        if new_pass:
            current_user.password = generate_password_hash(new_pass)
        db.session.commit()
        flash('Settings updated!', 'success')
        return redirect('/billing/settings')
    return render_template('settings.html')

@app.route('/billing/company', methods=['GET', 'POST'])
@login_required
def company_settings():
    company = CompanySettings.query.first()
    if not company:
        company = CompanySettings()
        db.session.add(company)
    if request.method == 'POST':
        company.company_name = request.form.get('company_name', company.company_name)
        company.tagline = request.form.get('tagline', company.tagline)
        company.address = request.form.get('address', company.address)
        company.city = request.form.get('city', company.city)
        company.state = request.form.get('state', company.state)
        company.pincode = request.form.get('pincode', company.pincode)
        company.phone = request.form.get('phone', company.phone)
        company.email = request.form.get('email', company.email)
        company.gstin = request.form.get('gstin', company.gstin)
        company.pan = request.form.get('pan', company.pan)
        company.bank_name = request.form.get('bank_name', company.bank_name)
        company.bank_account = request.form.get('bank_account', company.bank_account)
        company.bank_ifsc = request.form.get('bank_ifsc', company.bank_ifsc)
        db.session.commit()
        flash('Company settings updated!', 'success')
        return redirect('/billing/company')
    return render_template('company.html', company=company)

@app.route('/billing/reports/view')
@login_required
def view_report():
    today = date.today()
    now = datetime.now()
    filter_start = request.args.get('start_date', today.isoformat())
    filter_end = request.args.get('end_date', today.isoformat())
    start_dt = datetime.strptime(filter_start, '%Y-%m-%d')
    end_dt = datetime.strptime(filter_end, '%Y-%m-%d') + timedelta(days=1)

    today_start = datetime.combine(today, datetime.min.time())
    today_end = datetime.combine(today + timedelta(days=1), datetime.min.time())
    today_invoices_q = Invoice.query.filter(Invoice.date_created >= today_start, Invoice.date_created < today_end)
    today_invoices = today_invoices_q.count()
    today_sales = db.session.query(db.func.coalesce(db.func.sum(Invoice.grand_total), 0)).filter(Invoice.date_created >= today_start, Invoice.date_created < today_end).scalar()
    today_tax = db.session.query(db.func.coalesce(db.func.sum(Invoice.total_tax), 0)).filter(Invoice.date_created >= today_start, Invoice.date_created < today_end).scalar()
    today_paid_inv_ids = [p.invoice_id for p in Payment.query.filter(Payment.payment_date >= today_start, Payment.payment_date < today_end).all()]
    today_paid = db.session.query(db.func.coalesce(db.func.sum(Payment.amount), 0)).filter(Payment.payment_date >= today_start, Payment.payment_date < today_end).scalar()
    today_unpaid = today_sales - today_paid

    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    year_start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    total_invoices = Invoice.query.count()
    monthly_sales = db.session.query(db.func.coalesce(db.func.sum(Invoice.grand_total), 0)).filter(Invoice.date_created >= month_start).scalar()
    yearly_sales = db.session.query(db.func.coalesce(db.func.sum(Invoice.grand_total), 0)).filter(Invoice.date_created >= year_start).scalar()
    total_tax = db.session.query(db.func.coalesce(db.func.sum(Invoice.total_tax), 0)).scalar()

    filtered_invoices = Invoice.query.filter(Invoice.date_created >= start_dt, Invoice.date_created < end_dt).all()
    daily_map = {}
    for inv in filtered_invoices:
        day_str = inv.date_created.strftime('%Y-%m-%d')
        if day_str not in daily_map:
            daily_map[day_str] = {'date': day_str, 'invoices': 0, 'sales': 0, 'tax': 0, 'paid': 0, 'unpaid': 0}
        daily_map[day_str]['invoices'] += 1
        daily_map[day_str]['sales'] += inv.grand_total
        daily_map[day_str]['tax'] += inv.total_tax
        paid_amt = sum(p.amount for p in inv.payments)
        daily_map[day_str]['paid'] += paid_amt
        daily_map[day_str]['unpaid'] += (inv.grand_total - paid_amt)
    daily_data = sorted(daily_map.values(), key=lambda x: x['date'], reverse=True)

    product_sales_q = db.session.query(Product.name, db.func.sum(InvoiceItem.quantity), db.func.sum(InvoiceItem.total)).join(InvoiceItem, InvoiceItem.product_id == Product.id).join(Invoice, Invoice.id == InvoiceItem.invoice_id).filter(Invoice.date_created >= start_dt, Invoice.date_created < end_dt).group_by(Product.name).order_by(db.func.sum(InvoiceItem.quantity).desc()).limit(10).all()
    top_products = [{'name': n, 'total_qty': int(q or 0), 'total_revenue': float(r or 0)} for n, q, r in product_sales_q]

    product_data_q = db.session.query(Product.name, Product.stock, db.func.coalesce(db.func.sum(InvoiceItem.quantity), 0), db.func.coalesce(db.func.sum(InvoiceItem.total), 0)).outerjoin(InvoiceItem, InvoiceItem.product_id == Product.id).outerjoin(Invoice, Invoice.id == InvoiceItem.invoice_id).filter(Invoice.date_created >= start_dt, Invoice.date_created < end_dt).group_by(Product.id).all()
    product_data = [{'name': n, 'stock': s, 'sold': int(q or 0), 'revenue': float(r or 0)} for n, s, q, r in product_data_q]

    customer_data_q = db.session.query(Customer.name, Customer.phone, db.func.count(Invoice.id), db.func.coalesce(db.func.sum(Invoice.grand_total), 0)).outerjoin(Invoice, Invoice.customer_id == Customer.id).filter(Invoice.date_created >= start_dt, Invoice.date_created < end_dt).group_by(Customer.id).all()
    customer_data = [{'name': n, 'phone': p, 'invoices': int(c or 0), 'total': float(t or 0)} for n, p, c, t in customer_data_q]

    return render_template('view_report.html',
        today=now.strftime('%d-%m-%Y %H:%M'),
        filter_start=filter_start, filter_end=filter_end,
        today_invoices=today_invoices, today_sales=today_sales,
        today_tax=today_tax, today_paid=today_paid,
        total_invoices=total_invoices, monthly_sales=monthly_sales,
        yearly_sales=yearly_sales, total_tax=total_tax,
        daily_data=daily_data, top_products=top_products,
        product_data=product_data, customer_data=customer_data)

@app.route('/billing/invoice/<int:id>')
@login_required
def view_invoice(id):
    invoice = Invoice.query.get_or_404(id)
    company = CompanySettings.query.first()
    return render_template('invoice.html', invoice=invoice, company=company)

@app.route('/billing/invoice/<int:id>/print')
@login_required
def print_invoice(id):
    invoice = Invoice.query.get_or_404(id)
    company = CompanySettings.query.first()
    return render_template('print_invoice.html', invoice=invoice, company=company, now=datetime.now())

@app.route('/billing/<path:filename>')
def serve_billing_static(filename):
    billing_dir = os.path.join(BASE_DIR, 'Billing system')
    static_dir = os.path.join(billing_dir, 'static')
    if os.path.exists(os.path.join(static_dir, filename)):
        return send_from_directory(static_dir, filename)
    if os.path.exists(os.path.join(billing_dir, filename)):
        return send_from_directory(billing_dir, filename)
    return f"File not found: {filename}", 404


# ============================================
# Server-Side Auth API (portable users)
# ============================================
@app.route('/api/auth/register', methods=['POST'])
def api_auth_register():
    data = request.get_json()
    username = (data.get('username') or '').strip()
    password = data.get('password', '')
    role = data.get('role', 'staff').strip()
    full_name = (data.get('fullName') or data.get('full_name') or username).strip()
    if not username or not password:
        return jsonify({'success': False, 'message': 'Username and password required'}), 400
    if BUser.query.filter(BUser.username.ilike(username)).first():
        return jsonify({'success': False, 'message': 'Username already exists'}), 409
    user = BUser(username=username, password=generate_password_hash(password), full_name=full_name, role=role)
    db.session.add(user)
    db.session.commit()
    return jsonify({'success': True, 'user': {'id': user.id, 'username': user.username, 'role': user.role, 'fullName': user.full_name}})


@app.route('/api/auth/login', methods=['POST'])
def api_auth_login():
    data = request.get_json()
    username = (data.get('username') or '').strip()
    password = data.get('password', '')
    if not username or not password:
        return jsonify({'success': False, 'message': 'Username and password required'}), 400
    user = BUser.query.filter(BUser.username.ilike(username)).first()
    if not user or not check_password_hash(user.password, password):
        return jsonify({'success': False, 'message': 'Invalid username or password'}), 401
    return jsonify({'success': True, 'user': {'id': user.id, 'username': user.username, 'role': user.role, 'fullName': user.full_name}})


@app.route('/api/auth/users', methods=['GET'])
def api_auth_list_users():
    users = BUser.query.all()
    return jsonify({'success': True, 'users': [
        {'id': u.id, 'username': u.username, 'role': u.role, 'fullName': u.full_name}
        for u in users
    ]})


@app.route('/api/auth/users/<int:user_id>', methods=['DELETE'])
def api_auth_delete_user(user_id):
    user = BUser.query.get(user_id)
    if not user:
        return jsonify({'success': False, 'message': 'User not found'}), 404
    db.session.delete(user)
    db.session.commit()
    return jsonify({'success': True})


# ============================================
# Attendance System - Server-Side Storage
# ============================================
class PortalStaff(db.Model):
    __tablename__ = 'portal_staff'
    id = db.Column(db.Integer, primary_key=True)
    staff_id = db.Column(db.String(20), unique=True, nullable=False)
    full_name = db.Column(db.String(120), nullable=False)
    phone = db.Column(db.String(20), default='')
    location = db.Column(db.String(100), default='Office')
    job_title = db.Column(db.String(50), default='Staff')
    role = db.Column(db.String(20), default='staff')
    join_date = db.Column(db.String(20), default='')
    status = db.Column(db.String(20), default='active')

class AttRecord(db.Model):
    __tablename__ = 'att_record'
    id = db.Column(db.Integer, primary_key=True)
    staff_id = db.Column(db.String(20), nullable=False)
    date = db.Column(db.String(20), nullable=False)
    status = db.Column(db.String(20), default='present')
    clock_in = db.Column(db.String(10))
    clock_out = db.Column(db.String(10))
    hours = db.Column(db.Float, default=0)
    db.UniqueConstraint('staff_id', 'date')

class AttLeave(db.Model):
    __tablename__ = 'att_leave'
    id = db.Column(db.Integer, primary_key=True)
    leave_id = db.Column(db.String(20), unique=True, nullable=False)
    staff_id = db.Column(db.String(20), nullable=False)
    type = db.Column(db.String(30), default='Annual')
    start_date = db.Column(db.String(20))
    end_date = db.Column(db.String(20))
    reason = db.Column(db.Text, default='')
    status = db.Column(db.String(20), default='pending')
    applied_on = db.Column(db.String(20))

class AttTiming(db.Model):
    __tablename__ = 'att_timing'
    id = db.Column(db.Integer, primary_key=True)
    staff_id = db.Column(db.String(20), unique=True, nullable=False)
    start_time = db.Column(db.String(10), default='09:00')
    end_time = db.Column(db.String(10), default='18:00')
    grace_minutes = db.Column(db.Integer, default=10)
    half_day_hours = db.Column(db.Float, default=4)

class AttLog(db.Model):
    __tablename__ = 'att_log'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120))
    action = db.Column(db.String(30))
    time = db.Column(db.String(10))
    date = db.Column(db.String(20))
    timestamp = db.Column(db.BigInteger, default=0)

# ============================================
# Complaints System - Server-Side Storage
# ============================================
class Complaint(db.Model):
    __tablename__ = 'complaint'
    id = db.Column(db.Integer, primary_key=True)
    complaint_id = db.Column(db.BigInteger, unique=True, nullable=False)
    collection = db.Column(db.String(20), nullable=False, default='pending')
    customer_name = db.Column(db.String(200), default='')
    customer_phone = db.Column(db.String(20), default='')
    address = db.Column(db.Text, default='')
    city = db.Column(db.String(100), default='')
    location = db.Column(db.String(200), default='')
    complaint = db.Column(db.Text, default='')
    date = db.Column(db.String(20), default='')
    status = db.Column(db.String(20), default='Pending')
    created_at = db.Column(db.String(50), default='')
    started_at = db.Column(db.String(50), default='')
    completed_at = db.Column(db.String(50), default='')
    created_by = db.Column(db.String(100), default='')

with app.app_context():
    try:
        db.create_all()
    except Exception as e:
        print(f"[DB] Table creation error: {e}")

# --- Staff API ---
@app.route('/api/att/staff', methods=['GET'])
def api_att_get_staff():
    staff = PortalStaff.query.all()
    return jsonify({'success': True, 'data': [
        {'staffId': s.staff_id, 'fullName': s.full_name, 'phone': s.phone,
         'location': s.location, 'jobTitle': s.job_title, 'role': s.role,
         'joinDate': s.join_date, 'status': s.status}
        for s in staff
    ]})

@app.route('/api/att/staff', methods=['POST'])
def api_att_save_staff():
    data = request.get_json()
    items = data.get('data', data) if isinstance(data, dict) and 'data' in data else data
    if not isinstance(items, list):
        items = [items]
    for item in items:
        sid = item.get('staffId', '')
        if not sid: continue
        existing = PortalStaff.query.filter_by(staff_id=sid).first()
        if existing:
            existing.full_name = item.get('fullName', existing.full_name)
            existing.phone = item.get('phone', existing.phone)
            existing.location = item.get('location', existing.location)
            existing.job_title = item.get('jobTitle', existing.job_title)
            existing.role = item.get('role', existing.role)
            existing.join_date = item.get('joinDate', existing.join_date)
            existing.status = item.get('status', existing.status)
        else:
            db.session.add(PortalStaff(staff_id=sid, full_name=item.get('fullName', ''),
                phone=item.get('phone', ''), location=item.get('location', 'Office'),
                job_title=item.get('jobTitle', 'Staff'), role=item.get('role', 'staff'),
                join_date=item.get('joinDate', ''), status=item.get('status', 'active')))
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/att/staff/<staff_id>', methods=['DELETE'])
def api_att_delete_staff(staff_id):
    staff = PortalStaff.query.filter_by(staff_id=staff_id).first()
    if not staff:
        return jsonify({'success': False, 'message': 'Staff not found'}), 404
    AttRecord.query.filter_by(staff_id=staff_id).delete()
    AttLeave.query.filter_by(staff_id=staff_id).delete()
    AttTiming.query.filter_by(staff_id=staff_id).delete()
    db.session.delete(staff)
    db.session.commit()
    return jsonify({'success': True})

# --- Attendance Records API ---
@app.route('/api/att/records', methods=['GET'])
def api_att_get_records():
    date_f = request.args.get('date', '')
    staff_f = request.args.get('staffId', '')
    q = AttRecord.query
    if date_f: q = q.filter_by(date=date_f)
    if staff_f: q = q.filter_by(staff_id=staff_f)
    records = q.all()
    return jsonify({'success': True, 'data': [
        {'staffId': r.staff_id, 'date': r.date, 'status': r.status,
         'clockIn': r.clock_in, 'clockOut': r.clock_out, 'hours': r.hours}
        for r in records
    ]})

@app.route('/api/att/records', methods=['POST'])
def api_att_save_record():
    data = request.get_json()
    items = data.get('data', data) if isinstance(data, dict) and 'data' in data else data
    if not isinstance(items, list):
        items = [items]
    for item in items:
        sid = item.get('staffId', '')
        dt = item.get('date', '')
        if not sid or not dt: continue
        existing = AttRecord.query.filter_by(staff_id=sid, date=dt).first()
        if existing:
            existing.status = item.get('status', existing.status)
            existing.clock_in = item.get('clockIn', existing.clock_in)
            existing.clock_out = item.get('clockOut', existing.clock_out)
            existing.hours = item.get('hours', existing.hours)
        else:
            db.session.add(AttRecord(staff_id=sid, date=dt, status=item.get('status', 'present'),
                clock_in=item.get('clockIn'), clock_out=item.get('clockOut'), hours=item.get('hours', 0)))
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/att/records/<staff_id>/<date>', methods=['DELETE'])
def api_att_delete_record(staff_id, date):
    AttRecord.query.filter_by(staff_id=staff_id, date=date).delete()
    db.session.commit()
    return jsonify({'success': True})

# --- Leave API ---
@app.route('/api/att/leaves', methods=['GET'])
def api_att_get_leaves():
    leaves = AttLeave.query.all()
    return jsonify({'success': True, 'data': [
        {'id': l.leave_id, 'staffId': l.staff_id, 'type': l.type,
         'startDate': l.start_date, 'endDate': l.end_date, 'reason': l.reason,
         'status': l.status, 'appliedOn': l.applied_on}
        for l in leaves
    ]})

@app.route('/api/att/leaves', methods=['POST'])
def api_att_save_leave():
    data = request.get_json()
    items = data.get('data', data) if isinstance(data, dict) and 'data' in data else data
    if not isinstance(items, list):
        items = [items]
    for item in items:
        lid = item.get('id', '')
        if lid:
            existing = AttLeave.query.filter_by(leave_id=lid).first()
            if existing:
                existing.status = item.get('status', existing.status)
                existing.reason = item.get('reason', existing.reason)
                db.session.commit()
                continue
        max_id = db.session.query(db.func.max(AttLeave.id)).scalar() or 0
        lid = lid or f'LV{str(max_id + 1).zfill(3)}'
        db.session.add(AttLeave(leave_id=lid, staff_id=item.get('staffId', ''),
            type=item.get('type', 'Annual'), start_date=item.get('startDate', ''),
            end_date=item.get('endDate', ''), reason=item.get('reason', ''),
            status=item.get('status', 'pending'), applied_on=item.get('appliedOn', '')))
    db.session.commit()
    return jsonify({'success': True})

# --- Timings API ---
@app.route('/api/att/timings', methods=['GET'])
def api_att_get_timings():
    timings = AttTiming.query.all()
    result = {}
    for t in timings:
        result[t.staff_id] = {'startTime': t.start_time, 'endTime': t.end_time,
            'graceMinutes': t.grace_minutes, 'halfDayHours': t.half_day_hours}
    return jsonify({'success': True, 'data': result})

@app.route('/api/att/timings', methods=['POST'])
def api_att_save_timings():
    data = request.get_json()
    items = data.get('data', data) if isinstance(data, dict) and 'data' in data else data
    if not isinstance(items, dict):
        items = {}
    for sid, timings in items.items():
        existing = AttTiming.query.filter_by(staff_id=sid).first()
        if existing:
            existing.start_time = timings.get('startTime', existing.start_time)
            existing.end_time = timings.get('endTime', existing.end_time)
            existing.grace_minutes = timings.get('graceMinutes', existing.grace_minutes)
            existing.half_day_hours = timings.get('halfDayHours', existing.half_day_hours)
        else:
            db.session.add(AttTiming(staff_id=sid, start_time=timings.get('startTime', '09:00'),
                end_time=timings.get('endTime', '18:00'),
                grace_minutes=timings.get('graceMinutes', 10),
                half_day_hours=timings.get('halfDayHours', 4)))
    db.session.commit()
    return jsonify({'success': True})

# --- Recent Logs API ---
@app.route('/api/att/logs', methods=['GET'])
def api_att_get_logs():
    logs = AttLog.query.order_by(AttLog.timestamp.desc()).limit(20).all()
    return jsonify({'success': True, 'data': [
        {'name': l.name, 'action': l.action, 'time': l.time, 'date': l.date, 'timestamp': l.timestamp}
        for l in logs
    ]})

@app.route('/api/att/logs', methods=['POST'])
def api_att_save_log():
    data = request.get_json()
    items = data.get('data', data) if isinstance(data, dict) and 'data' in data else data
    if not isinstance(items, list):
        items = [items]
    for item in items:
        db.session.add(AttLog(name=item.get('name', ''), action=item.get('action', ''),
            time=item.get('time', ''), date=item.get('date', ''),
            timestamp=item.get('timestamp', 0)))
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/att/logs', methods=['DELETE'])
def api_att_clear_logs():
    AttLog.query.delete()
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/encrypt', methods=['POST'])
def api_encrypt():
    data = request.get_json()
    plaintext = data.get('data', '')
    password = data.get('password', '')
    try:
        encrypted = encrypt_data(plaintext, password) if password else encrypt_data(plaintext)
        return jsonify({'success': True, 'encrypted': encrypted})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/decrypt', methods=['POST'])
def api_decrypt():
    data = request.get_json()
    encrypted = data.get('encrypted', '')
    password = data.get('password', '')
    try:
        decrypted = decrypt_data(encrypted, password) if password else decrypt_data(encrypted)
        return jsonify({'success': True, 'data': decrypted})
    except Exception as e:
        return jsonify({'success': False, 'error': 'Decryption failed'}), 400

@app.route('/api/geo')
def get_client_geo():
    import urllib.request
    ip = request.headers.get('X-Forwarded-For', request.remote_addr)
    if ip and ip.startswith('127'):
        ip = ''
    try:
        url = f'http://ip-api.com/json/{ip}?fields=status,lat,lon,city,regionName,country'
        resp = urllib.request.urlopen(url, timeout=4)
        data = json.loads(resp.read())
        if data.get('status') == 'success':
            coords = f"{data['lat']},{data['lon']}"
            parts = [data.get('city'), data.get('regionName'), data.get('country')]
            address = ', '.join(p for p in parts if p)
            return jsonify({'success': True, 'coords': coords, 'address': address, 'lat': data['lat'], 'lon': data['lon']})
    except Exception:
        pass
    return jsonify({'success': False})

@app.route('/api/attendance/log', methods=['POST'])
def api_log_attendance():
    data = request.get_json()
    record = log_attendance(data.get('username', 'unknown'), data.get('system', 'portal'), data.get('role', 'user'))
    return jsonify({'success': True, 'record': record})

@app.route('/api/attendance/records')
def api_get_attendance():
    if attendance_log_file and os.path.exists(attendance_log_file):
        try:
            with open(attendance_log_file, 'r') as f:
                records = json.loads(decrypt_data(f.read()))
            date_f = request.args.get('date', '')
            system_f = request.args.get('system', '')
            user_f = request.args.get('user', '')
            if date_f:
                records = [r for r in records if r.get('date') == date_f]
            if system_f:
                records = [r for r in records if r.get('system') == system_f]
            if user_f:
                records = [r for r in records if r.get('username') == user_f]
            return jsonify({'success': True, 'records': records})
        except Exception:
            pass
    return jsonify({'success': True, 'records': []})

@app.route('/api/attendance/clear', methods=['POST'])
def api_clear_attendance():
    if attendance_log_file and os.path.exists(attendance_log_file):
        os.remove(attendance_log_file)
    return jsonify({'success': True})


# ============================================
# Firebase Admin SDK - Cloud Firestore Sync
# ============================================

# ============================================
# FILE UPLOAD CONFIGURATION - Single Server
# ============================================
USE_UPLOAD_SERVER = True
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
if not IS_VERCEL:
    os.makedirs(UPLOAD_DIR, exist_ok=True)

def decrypt_client_payload(data):
    if not data:
        return {}
    if isinstance(data, dict) and data.get('encrypted'):
        try:
            iv = base64.b64decode(data['iv'])
            cipher_bytes = base64.b64decode(data['data'])
            from cryptography.hazmat.primitives.ciphers.aead import AESGCM
            from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
            from cryptography.hazmat.primitives import hashes
            password = b'iacss_default_key_2026'
            salt = b'iacss_cloud_sync_v1'
            kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=100000)
            key = kdf.derive(password)
            aesgcm = AESGCM(key)
            plain = aesgcm.decrypt(iv, cipher_bytes, None)
            return json.loads(plain.decode('utf-8'))
        except Exception as e:
            print(f"[DECRYPT] Client payload failed: {e}")
            return data.get('data', {})
    if isinstance(data, dict) and data.get('plaintext'):
        return data.get('data', {})
    return data


# ============================================
# Cloud API: Employees
# ============================================
@app.route('/api/employees', methods=['GET'])
def api_cloud_employees():
    if not firestore_db:
        return jsonify({'success': True, 'data': []})
    try:
        docs = firestore_db.collection('employees').stream()
        employees = []
        for doc in docs:
            emp = doc.to_dict()
            emp['id'] = doc.id
            employees.append(emp)
        return jsonify({'success': True, 'data': employees})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/employees', methods=['POST'])
def api_cloud_create_employee():
    if not firestore_db:
        return jsonify({'success': False, 'error': 'Firestore not connected'}), 503
    try:
        data = decrypt_client_payload(request.get_json())
        doc_ref = firestore_db.collection('employees').add(data)
        return jsonify({'success': True, 'id': doc_ref[1].id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/employees/<emp_id>', methods=['PUT'])
def api_cloud_update_employee(emp_id):
    if not firestore_db:
        return jsonify({'success': False, 'error': 'Firestore not connected'}), 503
    try:
        data = decrypt_client_payload(request.get_json())
        firestore_db.collection('employees').document(emp_id).set(data, merge=True)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/employees/<emp_id>', methods=['DELETE'])
def api_cloud_delete_employee(emp_id):
    if not firestore_db:
        return jsonify({'success': False, 'error': 'Firestore not connected'}), 503
    try:
        firestore_db.collection('employees').document(emp_id).delete()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# Cloud API: Attendance
# ============================================
@app.route('/api/attendance', methods=['GET'])
def api_cloud_attendance():
    if not firestore_db:
        return jsonify({'success': True, 'data': []})
    try:
        docs = firestore_db.collection('attendance').order_by('date', direction=firestore.Query.DESCENDING).stream()
        records = []
        for doc in docs:
            rec = doc.to_dict()
            rec['id'] = doc.id
            records.append(rec)
        return jsonify({'success': True, 'data': records})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/attendance', methods=['POST'])
def api_cloud_create_attendance():
    if not firestore_db:
        return jsonify({'success': False, 'error': 'Firestore not connected'}), 503
    try:
        data = decrypt_client_payload(request.get_json())
        doc_ref = firestore_db.collection('attendance').add(data)
        return jsonify({'success': True, 'id': doc_ref[1].id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/attendance/<rec_id>', methods=['PUT'])
def api_cloud_update_attendance(rec_id):
    if not firestore_db:
        return jsonify({'success': False, 'error': 'Firestore not connected'}), 503
    try:
        data = decrypt_client_payload(request.get_json())
        firestore_db.collection('attendance').document(rec_id).set(data, merge=True)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/attendance/<rec_id>', methods=['DELETE'])
def api_cloud_delete_attendance(rec_id):
    if not firestore_db:
        return jsonify({'success': False, 'error': 'Firestore not connected'}), 503
    try:
        firestore_db.collection('attendance').document(rec_id).delete()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# Complaints API (SQLite-backed)
# ============================================
@app.route('/api/complaints/<collection>', methods=['GET'])
def api_complaints_list(collection):
    if collection not in ('pending', 'completed'):
        return jsonify({'success': False, 'error': 'Invalid collection'}), 400
    items = Complaint.query.filter_by(collection=collection).all()
    return jsonify({'success': True, 'data': [
        {
            'id': c.complaint_id, 'customerName': c.customer_name,
            'customerPhone': c.customer_phone, 'address': c.address,
            'city': c.city, 'location': c.location, 'complaint': c.complaint,
            'date': c.date, 'status': c.status, 'createdAt': c.created_at,
            'startedAt': c.started_at, 'completedAt': c.completed_at,
            'createdBy': c.created_by
        }
        for c in items
    ]})

def _extract_complaint_data(raw):
    if isinstance(raw, dict) and 'encrypted' in raw:
        return decrypt_client_payload(raw)
    if isinstance(raw, dict) and 'data' in raw and isinstance(raw['data'], dict):
        return raw['data']
    return raw

@app.route('/api/complaints/<collection>', methods=['POST'])
def api_complaints_create(collection):
    if collection not in ('pending', 'completed'):
        return jsonify({'success': False, 'error': 'Invalid collection'}), 400
    data = _extract_complaint_data(request.get_json())
    cid = data.get('id') or int(datetime.now().timestamp() * 1000)
    c = Complaint(
        complaint_id=cid, collection=collection,
        customer_name=data.get('customerName', ''),
        customer_phone=data.get('customerPhone', ''),
        address=data.get('address', ''),
        city=data.get('city', ''),
        location=data.get('location', ''),
        complaint=data.get('complaint', ''),
        date=data.get('date', ''),
        status=data.get('status', 'Pending'),
        created_at=data.get('createdAt', ''),
        started_at=data.get('startedAt', ''),
        completed_at=data.get('completedAt', ''),
        created_by=data.get('createdBy', '')
    )
    db.session.add(c)
    db.session.commit()
    return jsonify({'success': True, 'id': cid})

@app.route('/api/complaints/<collection>/<int:complaint_id>', methods=['PUT'])
def api_complaints_update(collection, complaint_id):
    if collection not in ('pending', 'completed'):
        return jsonify({'success': False, 'error': 'Invalid collection'}), 400
    c = Complaint.query.filter_by(complaint_id=complaint_id, collection=collection).first()
    if not c:
        return jsonify({'success': False, 'error': 'Not found'}), 404
    data = _extract_complaint_data(request.get_json())
    c.customer_name = data.get('customerName', c.customer_name)
    c.customer_phone = data.get('customerPhone', c.customer_phone)
    c.address = data.get('address', c.address)
    c.city = data.get('city', c.city)
    c.location = data.get('location', c.location)
    c.complaint = data.get('complaint', c.complaint)
    c.date = data.get('date', c.date)
    c.status = data.get('status', c.status)
    c.created_at = data.get('createdAt', c.created_at)
    c.started_at = data.get('startedAt', c.started_at)
    c.completed_at = data.get('completedAt', c.completed_at)
    c.created_by = data.get('createdBy', c.created_by)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/complaints/<collection>/<int:complaint_id>', methods=['DELETE'])
def api_complaints_delete(collection, complaint_id):
    if collection not in ('pending', 'completed'):
        return jsonify({'success': False, 'error': 'Invalid collection'}), 400
    Complaint.query.filter_by(complaint_id=complaint_id, collection=collection).delete()
    db.session.commit()
    return jsonify({'success': True})


# ============================================
# Cloud API: Users
# ============================================
@app.route('/api/users', methods=['GET'])
def api_cloud_users():
    if not firestore_db:
        return jsonify({'success': True, 'data': []})
    try:
        docs = firestore_db.collection('users').stream()
        users = []
        for doc in docs:
            u = doc.to_dict()
            u['id'] = doc.id
            users.append(u)
        return jsonify({'success': True, 'data': users})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/users', methods=['POST'])
def api_cloud_create_user():
    if not firestore_db:
        return jsonify({'success': False, 'error': 'Firestore not connected'}), 503
    try:
        data = decrypt_client_payload(request.get_json())
        doc_ref = firestore_db.collection('users').add(data)
        return jsonify({'success': True, 'id': doc_ref[1].id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/users/<user_id>', methods=['DELETE'])
def api_cloud_delete_user(user_id):
    if not firestore_db:
        return jsonify({'success': False, 'error': 'Firestore not connected'}), 503
    try:
        firestore_db.collection('users').document(user_id).delete()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# Cloud API: Settings (key-value store)
# ============================================
@app.route('/api/settings/<key>', methods=['GET'])
def api_cloud_get_setting(key):
    if not firestore_db:
        return jsonify({'success': True, 'data': None})
    try:
        doc = firestore_db.collection('settings').document(key).get()
        if doc.exists:
            val = doc.to_dict().get('value')
            print(f"[SYNC] Settings GET '{key}': type={type(val).__name__}, len={len(val) if isinstance(val, (list, dict, str)) else 'N/A'}")
            return jsonify({'success': True, 'data': val})
        print(f"[SYNC] Settings GET '{key}': not found")
        return jsonify({'success': True, 'data': None})
    except Exception as e:
        print(f"[SYNC] Settings GET '{key}' FAILED: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/settings/<key>', methods=['POST'])
def api_cloud_set_setting(key):
    if not firestore_db:
        return jsonify({'success': False, 'error': 'Firestore not connected'}), 503
    try:
        body = request.get_json()
        data = decrypt_client_payload(body)
        value = data.get('value', body.get('data', body))
        print(f"[SYNC] Settings SET '{key}': type={type(value).__name__}, len={len(value) if isinstance(value, (list, dict, str)) else 'N/A'}")
        firestore_db.collection('settings').document(key).set({'key': key, 'value': value}, merge=True)
        return jsonify({'success': True})
    except Exception as e:
        print(f"[SYNC] Settings SET '{key}' FAILED: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# Cloud API: Batch sync (bulk upload all local data)
# ============================================
@app.route('/api/sync/batch', methods=['POST'])
def api_cloud_batch_sync():
    if not firestore_db:
        return jsonify({'success': False, 'error': 'Firestore not connected'}), 503
    try:
        body = request.get_json()
        data = decrypt_client_payload(body)
        results = {}

        for collection_name in ['employees', 'attendance', 'users']:
            if collection_name in data and isinstance(data[collection_name], list):
                batch = firestore_db.batch()
                count = 0
                for item in data[collection_name]:
                    doc_id = item.get('id') or item.get('employeeId') or item.get('username')
                    if doc_id:
                        ref = firestore_db.collection(collection_name).document(str(doc_id))
                        batch.set(ref, item, merge=True)
                        count += 1
                        if count % 500 == 0:
                            batch.commit()
                            batch = firestore_db.batch()
                batch.commit()
                results[collection_name] = count

        for status in ['pending', 'completed']:
            key = 'complaints_' + status
            if key in data and isinstance(data[key], list):
                batch = firestore_db.batch()
                count = 0
                for item in data[key]:
                    doc_id = str(item.get('id', ''))
                    if doc_id:
                        ref = firestore_db.collection(key).document(doc_id)
                        batch.set(ref, item, merge=True)
                        count += 1
                batch.commit()
                results[key] = count

        if 'settings' in data and isinstance(data['settings'], dict):
            batch = firestore_db.batch()
            count = 0
            for skey, sval in data['settings'].items():
                ref = firestore_db.collection('settings').document(skey)
                batch.set(ref, {'key': skey, 'value': sval}, merge=True)
                count += 1
            batch.commit()
            results['settings'] = count

        return jsonify({'success': True, 'synced': results})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# Cloud API: Full pull (get all data at once)
# ============================================
@app.route('/api/sync/pull', methods=['GET'])
def api_cloud_full_pull():
    if not firestore_db:
        return jsonify({'success': True, 'data': {}})
    try:
        result = {}
        for collection_name in ['employees', 'attendance', 'users']:
            docs = firestore_db.collection(collection_name).stream()
            result[collection_name] = [{'id': d.id, **d.to_dict()} for d in docs]

        for status in ['pending', 'completed']:
            key = 'complaints_' + status
            docs = firestore_db.collection(key).stream()
            result[key] = [{'id': d.id, **d.to_dict()} for d in docs]

        settings_docs = firestore_db.collection('settings').stream()
        settings = {}
        for d in settings_docs:
            data = d.to_dict()
            settings[d.id] = data.get('value') if data else None
        result['settings'] = settings

        return jsonify({'success': True, 'data': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# File Upload/Download with Supabase Storage
# ============================================
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
if not IS_VERCEL:
    os.makedirs(UPLOAD_DIR, exist_ok=True)

_supabase_storage = None

def get_supabase_storage():
    global _supabase_storage
    if _supabase_storage is None and SUPABASE_URL and SUPABASE_KEY:
        try:
            from supabase import create_client
            client = create_client(SUPABASE_URL, SUPABASE_KEY)
            _supabase_storage = client.storage
            try:
                _supabase_storage.get_bucket(SUPABASE_STORAGE_BUCKET)
            except Exception:
                _supabase_storage.create_bucket(SUPABASE_STORAGE_BUCKET, public=True)
        except Exception as e:
            print(f"[Storage] Supabase init failed: {e}")
    return _supabase_storage

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'success': False, 'error': 'No filename'}), 400

    safe_name = f"{int(datetime.now().timestamp() * 1000)}_{f.filename.replace(' ', '_')}"

    storage = get_supabase_storage()
    if storage:
        try:
            file_bytes = f.read()
            content_type = mimetypes.guess_type(f.filename)[0] or 'application/octet-stream'
            storage.from_(SUPABASE_STORAGE_BUCKET).upload(safe_name, file_bytes, {'content-type': content_type})
            public_url = storage.from_(SUPABASE_STORAGE_BUCKET).get_public_url(safe_name)
            print(f"[Storage] Uploaded to Supabase: {public_url}")
            return jsonify({'success': True, 'filename': safe_name, 'original': f.filename, 'url': public_url})
        except Exception as e:
            print(f"[Storage] Supabase upload failed, falling back to local: {e}")
            f.seek(0)

    f.save(os.path.join(UPLOAD_DIR, safe_name))
    return jsonify({'success': True, 'filename': safe_name, 'original': f.filename})

@app.route('/uploads/<path:filename>')
def serve_upload(filename):
    storage = get_supabase_storage()
    if storage:
        try:
            public_url = storage.from_(SUPABASE_STORAGE_BUCKET).get_public_url(filename)
            return redirect(public_url)
        except Exception:
            pass
    return send_from_directory(UPLOAD_DIR, filename)


def auto_migrate():
    try:
        from sqlalchemy import inspect as sa_inspect
        inspector = sa_inspect(db.engine)
        migrations = []
        table_columns = {
            'product': ['price_with_gst', 'brand', 'category'],
            'invoice_item': ['star_rating', 'model_no', 'series_no', 'item_category', 'ton_size'],
            'invoice': ['reverse_charge'],
        }
        for table, cols in table_columns.items():
            existing = [c['name'] for c in inspector.get_columns(table)]
            for col in cols:
                if col not in existing:
                    col_type = 'VARCHAR(100)' if col != 'price_with_gst' else 'FLOAT DEFAULT 0'
                    db.session.execute(db.text(f'ALTER TABLE {table} ADD COLUMN {col} {col_type}'))
                    migrations.append(f'{table}.{col}')
        if migrations:
            db.session.commit()
            print(f"[MIGRATE] Added columns: {', '.join(migrations)}")
    except Exception as e:
        print(f"[MIGRATE] Warning: {e}")

with app.app_context():
    try:
        db.create_all()
        auto_migrate()
        if not BUser.query.filter_by(username='Ramesh').first():
            db.session.add(BUser(username='Ramesh', password=generate_password_hash('Indiana/c'), full_name='Ramesh', role='admin'))
        if not BUser.query.filter_by(username='staff').first():
            db.session.add(BUser(username='staff', password=generate_password_hash('staff123'), full_name='Staff Member', role='staff'))
        db.session.commit()
    except Exception as e:
        print(f"[DB] Init error: {e}")


if __name__ == '__main__':
    print("=" * 60)
    print("  INDIAN A/C SALES & SERVICES - Unified Encrypted Server")
    print("=" * 60)
    print(f"  Portal:     http://localhost:5000/")
    print(f"  Attendance: http://localhost:5000/attendance/")
    print(f"  Complaints: http://localhost:5000/complaints/")
    print(f"  Billing:    http://localhost:5000/billing/")
    print(f"  Employees:  http://localhost:5000/employees/")
    print(f"  Logs:       http://localhost:5000/attendance-log.html")
    print(f"  Encryption: AES-256 (Fernet)")
    print("=" * 60)
    print(f"  Portal login: Ramesh / Indiana/c")
    print(f"  Billing login: Ramesh / Indiana/c, staff / staff123")
    print("=" * 60)
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True, threaded=True)
